"""Financial report generation: P&L, Balance Sheet, General Ledger.

Reports are computed from verification data and can be exported to Excel.
"""

from decimal import Decimal
from collections import OrderedDict
from io import BytesIO
from sqlalchemy import func

from app.extensions import db
from app.models.accounting import Account, Verification, VerificationRow, FiscalYear
from app.models.company import Company


def get_profit_and_loss(company_id, fiscal_year_id):
    """Generate Resultaträkning (P&L / Income Statement).

    Groups accounts by class (3xxx-8xxx) and calculates totals.
    Revenue accounts (3xxx) are shown as positive (credit - debit).
    Expense accounts (4xxx-7xxx) are shown as positive (debit - credit).
    """
    results = _get_account_balances(company_id, fiscal_year_id)

    sections = OrderedDict()

    # Revenue (3xxx) - credit balances are positive
    revenue_accounts = [(a, -b) for a, b in results if a.account_number[0] == '3']
    sections['Nettoomsättning'] = {
        'accounts': revenue_accounts,
        'total': sum(b for _, b in revenue_accounts),
    }

    # Cost of goods (4xxx)
    cogs_accounts = [(a, b) for a, b in results if a.account_number[0] == '4']
    sections['Kostnad sålda varor'] = {
        'accounts': cogs_accounts,
        'total': sum(b for _, b in cogs_accounts),
    }

    gross_profit = sections['Nettoomsättning']['total'] - sections['Kostnad sålda varor']['total']

    # External costs (5xxx-6xxx)
    ext_accounts = [(a, b) for a, b in results if a.account_number[0] in ('5', '6')]
    sections['Övriga externa kostnader'] = {
        'accounts': ext_accounts,
        'total': sum(b for _, b in ext_accounts),
    }

    # Personnel costs (7xxx)
    personnel_accounts = [(a, b) for a, b in results if a.account_number[0] == '7']
    sections['Personalkostnader'] = {
        'accounts': personnel_accounts,
        'total': sum(b for _, b in personnel_accounts),
    }

    operating_result = gross_profit - (
        sections['Övriga externa kostnader']['total'] +
        sections['Personalkostnader']['total']
    )

    # Financial items (8xxx)
    fin_revenue = [(a, -b) for a, b in results
                   if a.account_number[0] == '8' and a.account_type == 'revenue']
    fin_expense = [(a, b) for a, b in results
                   if a.account_number[0] == '8' and a.account_type == 'expense']

    sections['Finansiella intäkter'] = {
        'accounts': fin_revenue,
        'total': sum(b for _, b in fin_revenue),
    }
    sections['Finansiella kostnader'] = {
        'accounts': fin_expense,
        'total': sum(b for _, b in fin_expense),
    }

    result_before_tax = operating_result + (
        sections['Finansiella intäkter']['total'] -
        sections['Finansiella kostnader']['total']
    )

    return {
        'sections': sections,
        'gross_profit': float(gross_profit),
        'operating_result': float(operating_result),
        'result_before_tax': float(result_before_tax),
    }


def get_balance_sheet(company_id, fiscal_year_id):
    """Generate Balansräkning (Balance Sheet).

    Groups accounts by class (1xxx for assets, 2xxx for equity+liabilities).
    """
    results = _get_account_balances(company_id, fiscal_year_id)

    sections = OrderedDict()

    # Assets (1xxx) - debit balances
    fixed_assets = [(a, b) for a, b in results
                    if a.account_number[:2] in ('10', '11', '12', '13')]
    sections['Anläggningstillgångar'] = {
        'accounts': fixed_assets,
        'total': sum(b for _, b in fixed_assets),
    }

    current_assets = [(a, b) for a, b in results
                      if a.account_number[0] == '1' and a.account_number[:2] not in ('10', '11', '12', '13')]
    sections['Omsättningstillgångar'] = {
        'accounts': current_assets,
        'total': sum(b for _, b in current_assets),
    }

    total_assets = sections['Anläggningstillgångar']['total'] + sections['Omsättningstillgångar']['total']

    # Equity (20xx)
    equity_accounts = [(a, -b) for a, b in results if a.account_number[:2] in ('20', '21')]
    sections['Eget kapital'] = {
        'accounts': equity_accounts,
        'total': sum(b for _, b in equity_accounts),
    }

    # Long-term liabilities (23xx)
    lt_liabilities = [(a, -b) for a, b in results if a.account_number[:2] in ('22', '23')]
    sections['Långfristiga skulder'] = {
        'accounts': lt_liabilities,
        'total': sum(b for _, b in lt_liabilities),
    }

    # Short-term liabilities (24xx-29xx)
    st_liabilities = [(a, -b) for a, b in results
                      if a.account_number[0] == '2' and a.account_number[:2] not in ('20', '21', '22', '23')]
    sections['Kortfristiga skulder'] = {
        'accounts': st_liabilities,
        'total': sum(b for _, b in st_liabilities),
    }

    total_equity_liabilities = (
        sections['Eget kapital']['total'] +
        sections['Långfristiga skulder']['total'] +
        sections['Kortfristiga skulder']['total']
    )

    return {
        'sections': sections,
        'total_assets': float(total_assets),
        'total_equity_liabilities': float(total_equity_liabilities),
    }


def get_general_ledger(company_id, fiscal_year_id, account_number=None):
    """Generate Huvudbok (General Ledger).

    Shows all transactions per account, optionally filtered to one account.
    """
    query = db.session.query(
        Account,
        VerificationRow,
        Verification,
    ).join(
        VerificationRow, VerificationRow.account_id == Account.id
    ).join(
        Verification, Verification.id == VerificationRow.verification_id
    ).filter(
        Account.company_id == company_id,
        Verification.fiscal_year_id == fiscal_year_id,
    )

    if account_number:
        query = query.filter(Account.account_number == account_number)

    query = query.order_by(Account.account_number, Verification.verification_date)

    results = query.all()

    ledger = OrderedDict()
    for account, row, ver in results:
        key = f'{account.account_number} {account.name}'
        if key not in ledger:
            ledger[key] = {
                'account_number': account.account_number,
                'account_name': account.name,
                'entries': [],
                'total_debit': Decimal('0'),
                'total_credit': Decimal('0'),
            }

        ledger[key]['entries'].append({
            'date': ver.verification_date,
            'verification_number': ver.verification_number,
            'description': ver.description or row.description or '',
            'debit': float(row.debit),
            'credit': float(row.credit),
        })
        ledger[key]['total_debit'] += row.debit
        ledger[key]['total_credit'] += row.credit

    # Convert Decimals
    for key in ledger:
        ledger[key]['total_debit'] = float(ledger[key]['total_debit'])
        ledger[key]['total_credit'] = float(ledger[key]['total_credit'])
        ledger[key]['balance'] = ledger[key]['total_debit'] - ledger[key]['total_credit']

    return ledger


def export_report_to_excel(report_data, report_type, company_name, fiscal_year):
    """Export a report to Excel format.

    Args:
        report_data: Report dict from get_profit_and_loss/get_balance_sheet/get_general_ledger
        report_type: 'pnl', 'balance', 'ledger'
        company_name: Company name for the header
        fiscal_year: FiscalYear object

    Returns:
        BytesIO with Excel file content
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active

    bold = Font(bold=True)
    header_font = Font(bold=True, size=14)
    thin_border = Border(bottom=Side(style='thin'))

    if report_type == 'pnl':
        ws.title = 'Resultaträkning'
        ws.append([company_name])
        ws['A1'].font = header_font
        ws.append([f'Resultaträkning {fiscal_year.start_date} - {fiscal_year.end_date}'])
        ws.append([])

        for section_name, section in report_data['sections'].items():
            ws.append([section_name])
            ws.cell(row=ws.max_row, column=1).font = bold
            for account, balance in section['accounts']:
                if abs(balance) > 0.01:
                    ws.append([f'  {account.account_number}', account.name, '', round(balance, 2)])
            ws.append(['', f'Summa {section_name}', '', round(section['total'], 2)])
            ws.cell(row=ws.max_row, column=2).font = bold
            ws.append([])

        ws.append(['Rörelseresultat', '', '', round(report_data['operating_result'], 2)])
        ws.cell(row=ws.max_row, column=1).font = bold
        ws.append(['Resultat före skatt', '', '', round(report_data['result_before_tax'], 2)])
        ws.cell(row=ws.max_row, column=1).font = bold

    elif report_type == 'balance':
        ws.title = 'Balansräkning'
        ws.append([company_name])
        ws['A1'].font = header_font
        ws.append([f'Balansräkning per {fiscal_year.end_date}'])
        ws.append([])

        ws.append(['TILLGÅNGAR'])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        for section_name in ('Anläggningstillgångar', 'Omsättningstillgångar'):
            section = report_data['sections'][section_name]
            ws.append([section_name])
            ws.cell(row=ws.max_row, column=1).font = bold
            for account, balance in section['accounts']:
                if abs(balance) > 0.01:
                    ws.append([f'  {account.account_number}', account.name, '', round(balance, 2)])
            ws.append(['', f'Summa {section_name}', '', round(section['total'], 2)])
            ws.cell(row=ws.max_row, column=2).font = bold

        ws.append(['Summa tillgångar', '', '', round(report_data['total_assets'], 2)])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
        ws.append([])

        ws.append(['EGET KAPITAL OCH SKULDER'])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        for section_name in ('Eget kapital', 'Långfristiga skulder', 'Kortfristiga skulder'):
            section = report_data['sections'][section_name]
            ws.append([section_name])
            ws.cell(row=ws.max_row, column=1).font = bold
            for account, balance in section['accounts']:
                if abs(balance) > 0.01:
                    ws.append([f'  {account.account_number}', account.name, '', round(balance, 2)])
            ws.append(['', f'Summa {section_name}', '', round(section['total'], 2)])
            ws.cell(row=ws.max_row, column=2).font = bold

        ws.append(['Summa eget kapital och skulder', '', '', round(report_data['total_equity_liabilities'], 2)])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

    elif report_type == 'ledger':
        ws.title = 'Huvudbok'
        ws.append([company_name])
        ws['A1'].font = header_font
        ws.append([f'Huvudbok {fiscal_year.start_date} - {fiscal_year.end_date}'])
        ws.append([])

        for key, data in report_data.items():
            ws.append([f'{data["account_number"]} {data["account_name"]}'])
            ws.cell(row=ws.max_row, column=1).font = bold
            ws.append(['Datum', 'Ver.nr', 'Beskrivning', 'Debet', 'Kredit'])
            for cell in ws[ws.max_row]:
                cell.font = bold

            for entry in data['entries']:
                ws.append([
                    entry['date'].isoformat() if entry['date'] else '',
                    entry['verification_number'],
                    entry['description'],
                    round(entry['debit'], 2) if entry['debit'] else '',
                    round(entry['credit'], 2) if entry['credit'] else '',
                ])

            ws.append(['', '', 'Saldo',
                        round(data['total_debit'], 2),
                        round(data['total_credit'], 2)])
            ws.cell(row=ws.max_row, column=3).font = bold
            ws.append([])

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _get_account_balances(company_id, fiscal_year_id):
    """Get all account balances for a fiscal year.

    Returns list of (Account, balance) tuples.
    Balance = total_debit - total_credit.
    """
    results = db.session.query(
        Account,
        func.coalesce(func.sum(VerificationRow.debit), 0).label('total_debit'),
        func.coalesce(func.sum(VerificationRow.credit), 0).label('total_credit'),
    ).join(
        VerificationRow, VerificationRow.account_id == Account.id
    ).join(
        Verification, Verification.id == VerificationRow.verification_id
    ).filter(
        Account.company_id == company_id,
        Verification.fiscal_year_id == fiscal_year_id,
    ).group_by(
        Account.id
    ).order_by(
        Account.account_number
    ).all()

    return [(r[0], float(r[1] - r[2])) for r in results]
