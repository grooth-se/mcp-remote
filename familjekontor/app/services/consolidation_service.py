"""Multi-company consolidation service: consolidated P&L, balance sheet,
intercompany detection, minority interest, goodwill, cash flow."""

from decimal import Decimal
from collections import OrderedDict
from datetime import date
from io import BytesIO

from app.extensions import db
from app.models.consolidation import (
    ConsolidationGroup, ConsolidationGroupMember, IntercompanyElimination,
    IntercompanyMatch, AcquisitionGoodwill,
)
from app.models.accounting import FiscalYear, Account, Verification, VerificationRow
from app.models.company import Company
from app.services.report_service import get_profit_and_loss, get_balance_sheet
from sqlalchemy import func


# ---------------------------------------------------------------------------
# Group & Member CRUD
# ---------------------------------------------------------------------------

def create_consolidation_group(name, parent_company_id, description=None):
    group = ConsolidationGroup(
        name=name,
        parent_company_id=parent_company_id,
        description=description,
    )
    db.session.add(group)
    db.session.commit()
    return group


def add_member(group_id, company_id, ownership_pct=100, consolidation_method='full',
               parent_member_id=None):
    existing = ConsolidationGroupMember.query.filter_by(
        group_id=group_id, company_id=company_id
    ).first()
    if existing:
        existing.ownership_pct = ownership_pct
        existing.consolidation_method = consolidation_method
        existing.parent_member_id = parent_member_id
        db.session.commit()
        return existing

    member = ConsolidationGroupMember(
        group_id=group_id,
        company_id=company_id,
        ownership_pct=ownership_pct,
        consolidation_method=consolidation_method,
        parent_member_id=parent_member_id,
    )
    db.session.add(member)
    db.session.commit()
    return member


def remove_member(group_id, company_id):
    member = ConsolidationGroupMember.query.filter_by(
        group_id=group_id, company_id=company_id
    ).first()
    if member:
        db.session.delete(member)
        db.session.commit()
        return True
    return False


def _get_fy_for_company(company_id, year):
    """Get fiscal year for a company matching a specific year."""
    return FiscalYear.query.filter_by(
        company_id=company_id, year=year
    ).first()


# ---------------------------------------------------------------------------
# Minority interest
# ---------------------------------------------------------------------------

def calculate_effective_ownership(group_id, company_id):
    """Walk ownership chain to get effective ownership percentage."""
    member = ConsolidationGroupMember.query.filter_by(
        group_id=group_id, company_id=company_id
    ).first()
    if not member:
        return Decimal('0')

    pct = Decimal(str(member.ownership_pct or 100))
    current = member
    while current.parent_member_id:
        parent = db.session.get(ConsolidationGroupMember, current.parent_member_id)
        if not parent:
            break
        pct = pct * Decimal(str(parent.ownership_pct or 100)) / Decimal('100')
        current = parent

    return pct


def calculate_minority_interest(group_id, fy_year):
    """Calculate minority interest for each subsidiary with < 100% ownership.

    Returns list of dicts with company, ownership_pct, minority_pct,
    minority_pnl (share of P&L), minority_equity (share of equity).
    """
    group = db.session.get(ConsolidationGroup, group_id)
    if not group:
        return []

    result = []
    for member in group.members:
        if member.consolidation_method != 'full':
            continue
        effective_pct = float(calculate_effective_ownership(group_id, member.company_id))
        if effective_pct >= 100:
            continue

        minority_pct = 100.0 - effective_pct
        fy = _get_fy_for_company(member.company_id, fy_year)
        if not fy:
            continue

        pnl = get_profit_and_loss(member.company_id, fy.id)
        bs = get_balance_sheet(member.company_id, fy.id)

        pnl_result = pnl.get('result_before_tax', 0)
        equity_total = bs['sections'].get('Eget kapital', {}).get('total', 0)

        result.append({
            'company': member.company,
            'ownership_pct': effective_pct,
            'minority_pct': minority_pct,
            'minority_pnl': float(pnl_result) * minority_pct / 100.0,
            'minority_equity': float(equity_total) * minority_pct / 100.0,
        })

    return result


# ---------------------------------------------------------------------------
# Consolidated P&L
# ---------------------------------------------------------------------------

def get_consolidated_pnl(group_id, fy_year):
    """Generate consolidated P&L across all group members."""
    group = db.session.get(ConsolidationGroup, group_id)
    if not group:
        return None

    member_data = []
    consolidated_sections = OrderedDict()

    for member in group.members:
        fy = _get_fy_for_company(member.company_id, fy_year)
        if not fy:
            continue

        pnl = get_profit_and_loss(member.company_id, fy.id)
        method = member.consolidation_method or 'full'

        if method == 'full':
            weight = 1.0  # Full consolidation: 100% of all items
        elif method == 'equity':
            weight = float(member.ownership_pct or 100) / 100.0
        else:
            weight = 0.0  # cost method: no P&L consolidation

        member_data.append({
            'company': member.company,
            'ownership_pct': float(member.ownership_pct or 100),
            'consolidation_method': method,
            'pnl': pnl,
            'weight': weight,
        })

        for section_name, section in pnl['sections'].items():
            if section_name not in consolidated_sections:
                consolidated_sections[section_name] = {'total': 0, 'accounts': []}
            weighted_total = float(section['total']) * weight
            consolidated_sections[section_name]['total'] += weighted_total

    gross_profit = consolidated_sections.get('Nettoomsättning', {}).get('total', 0) - \
                   consolidated_sections.get('Kostnad sålda varor', {}).get('total', 0)

    operating_result = gross_profit - (
        consolidated_sections.get('Övriga externa kostnader', {}).get('total', 0) +
        consolidated_sections.get('Personalkostnader', {}).get('total', 0)
    )

    result_before_tax = operating_result + (
        consolidated_sections.get('Finansiella intäkter', {}).get('total', 0) -
        consolidated_sections.get('Finansiella kostnader', {}).get('total', 0)
    )

    # Apply eliminations
    eliminations = IntercompanyElimination.query.filter_by(group_id=group_id).all()
    elimination_total = sum(float(e.amount) for e in eliminations)

    # Goodwill amortization for the year
    goodwill_total = _get_goodwill_amortization_for_year(group_id, fy_year)

    # Minority interest
    minority = calculate_minority_interest(group_id, fy_year)
    minority_pnl_total = sum(m['minority_pnl'] for m in minority)

    return {
        'group': group,
        'members': member_data,
        'sections': consolidated_sections,
        'gross_profit': gross_profit,
        'operating_result': operating_result,
        'result_before_tax': result_before_tax,
        'elimination_total': elimination_total,
        'goodwill_amortization': goodwill_total,
        'minority_pnl': minority_pnl_total,
        'minority_details': minority,
        'adjusted_result': result_before_tax - elimination_total - goodwill_total - minority_pnl_total,
        'report_type': 'pnl',
    }


# ---------------------------------------------------------------------------
# Consolidated Balance Sheet
# ---------------------------------------------------------------------------

def get_consolidated_balance_sheet(group_id, fy_year):
    """Generate consolidated balance sheet across all group members."""
    group = db.session.get(ConsolidationGroup, group_id)
    if not group:
        return None

    member_data = []
    consolidated_sections = OrderedDict()

    for member in group.members:
        fy = _get_fy_for_company(member.company_id, fy_year)
        if not fy:
            continue

        bs = get_balance_sheet(member.company_id, fy.id)
        method = member.consolidation_method or 'full'

        if method == 'full':
            weight = 1.0
        elif method == 'equity':
            weight = float(member.ownership_pct or 100) / 100.0
        else:
            weight = 0.0

        member_data.append({
            'company': member.company,
            'ownership_pct': float(member.ownership_pct or 100),
            'consolidation_method': method,
            'balance_sheet': bs,
            'weight': weight,
        })

        for section_name, section in bs['sections'].items():
            if section_name not in consolidated_sections:
                consolidated_sections[section_name] = {'total': 0}
            weighted_total = float(section['total']) * weight
            consolidated_sections[section_name]['total'] += weighted_total

    # Add goodwill to fixed assets
    remaining_goodwill = get_total_remaining_goodwill(group_id)
    if remaining_goodwill > 0:
        if 'Anläggningstillgångar' not in consolidated_sections:
            consolidated_sections['Anläggningstillgångar'] = {'total': 0}
        consolidated_sections['Anläggningstillgångar']['total'] += remaining_goodwill

    total_assets = (
        consolidated_sections.get('Anläggningstillgångar', {}).get('total', 0) +
        consolidated_sections.get('Omsättningstillgångar', {}).get('total', 0)
    )

    # Minority interest in equity
    minority = calculate_minority_interest(group_id, fy_year)
    minority_equity_total = sum(m['minority_equity'] for m in minority)

    total_equity_liabilities = (
        consolidated_sections.get('Eget kapital', {}).get('total', 0) +
        consolidated_sections.get('Långfristiga skulder', {}).get('total', 0) +
        consolidated_sections.get('Kortfristiga skulder', {}).get('total', 0)
    )

    return {
        'group': group,
        'members': member_data,
        'sections': consolidated_sections,
        'total_assets': total_assets,
        'total_equity_liabilities': total_equity_liabilities,
        'remaining_goodwill': remaining_goodwill,
        'minority_equity': minority_equity_total,
        'minority_details': minority,
        'report_type': 'balance',
    }


# ---------------------------------------------------------------------------
# Eliminations
# ---------------------------------------------------------------------------

def create_elimination(group_id, fiscal_year_id, from_company_id, to_company_id,
                       account_number, amount, description=None):
    elimination = IntercompanyElimination(
        group_id=group_id,
        fiscal_year_id=fiscal_year_id,
        from_company_id=from_company_id,
        to_company_id=to_company_id,
        account_number=account_number,
        amount=amount,
        description=description,
    )
    db.session.add(elimination)
    db.session.commit()
    return elimination


# ---------------------------------------------------------------------------
# Intercompany Detection
# ---------------------------------------------------------------------------

def detect_intercompany_transactions(group_id, fiscal_year_id):
    """Scan group members for matching counterparty transactions.

    Looks for invoice amounts that match between two group companies
    (one as supplier invoice, one as customer invoice for same amount).
    Also looks for matching balances on 1660 (receivables) / 2360 (payables).
    Returns list of suggested IntercompanyMatch objects.
    """
    group = db.session.get(ConsolidationGroup, group_id)
    if not group:
        return []

    member_company_ids = [m.company_id for m in group.members]
    if len(member_company_ids) < 2:
        return []

    fy = db.session.get(FiscalYear, fiscal_year_id)
    if not fy:
        return []

    matches_found = []

    # Strategy: for each pair of companies, look for matching amounts in
    # their receivable (1660) and payable (2360) accounts
    # Each company may have its own fiscal year ID for the same year
    fy_map = {}
    for member in group.members:
        member_fy = FiscalYear.query.filter_by(
            company_id=member.company_id, year=fy.year
        ).first()
        if member_fy:
            fy_map[member.company_id] = member_fy.id

    for i, company_a_id in enumerate(member_company_ids):
        for company_b_id in member_company_ids[i + 1:]:
            fy_a = fy_map.get(company_a_id)
            fy_b = fy_map.get(company_b_id)
            if not fy_a or not fy_b:
                continue
            _find_matching_balances(
                group_id, fy_a, fy_b,
                company_a_id, company_b_id,
                matches_found,
            )

    # Persist suggested matches
    created = []
    for match_data in matches_found:
        # Check for existing match
        existing = IntercompanyMatch.query.filter_by(
            group_id=group_id,
            fiscal_year_id=fiscal_year_id,
            company_a_id=match_data['company_a_id'],
            company_b_id=match_data['company_b_id'],
            amount=match_data['amount'],
        ).first()
        if existing:
            continue

        match = IntercompanyMatch(
            group_id=group_id,
            fiscal_year_id=fiscal_year_id,
            company_a_id=match_data['company_a_id'],
            company_b_id=match_data['company_b_id'],
            amount=match_data['amount'],
            match_type=match_data.get('match_type', 'invoice'),
            description=match_data.get('description', ''),
            status='suggested',
        )
        db.session.add(match)
        created.append(match)

    if created:
        db.session.commit()
    return created


def _find_matching_balances(group_id, fy_a_id, fy_b_id, company_a_id, company_b_id,
                            matches_found):
    """Find matching receivable/payable balances between two companies."""
    receivable_accts = ['1660', '1661', '1662']
    payable_accts = ['2360', '2361', '2362']

    a_receivable = _sum_account_balances(company_a_id, fy_a_id, receivable_accts)
    b_payable = _sum_account_balances(company_b_id, fy_b_id, payable_accts)

    # If A has receivables and B has matching payables, suggest elimination
    if a_receivable > 0 and b_payable > 0:
        match_amount = min(a_receivable, b_payable)
        if match_amount > Decimal('0.01'):
            matches_found.append({
                'company_a_id': company_a_id,
                'company_b_id': company_b_id,
                'amount': match_amount,
                'match_type': 'loan',
                'description': f'Koncernintern fordran/skuld: {match_amount}',
            })

    # Check reverse direction
    b_receivable = _sum_account_balances(company_b_id, fy_b_id, receivable_accts)
    a_payable = _sum_account_balances(company_a_id, fy_a_id, payable_accts)

    if b_receivable > 0 and a_payable > 0:
        match_amount = min(b_receivable, a_payable)
        if match_amount > Decimal('0.01'):
            matches_found.append({
                'company_a_id': company_b_id,
                'company_b_id': company_a_id,
                'amount': match_amount,
                'match_type': 'loan',
                'description': f'Koncernintern fordran/skuld: {match_amount}',
            })

    # Check intercompany sales (3xxx) vs purchases (4xxx)
    a_sales = _sum_account_range(company_a_id, fy_a_id, '3')
    b_purchases = _sum_account_range(company_b_id, fy_b_id, '4')

    if a_sales > 0 and b_purchases > 0:
        # Heuristic: if A sells and B buys, some may be intercompany
        # Use the smaller amount as potential match
        potential = min(a_sales, b_purchases)
        if potential > Decimal('100'):  # Only suggest significant amounts
            matches_found.append({
                'company_a_id': company_a_id,
                'company_b_id': company_b_id,
                'amount': potential,
                'match_type': 'invoice',
                'description': f'Potentiell koncernintern försäljning/inköp: {potential}',
            })


def _sum_account_balances(company_id, fiscal_year_id, account_numbers):
    """Sum absolute balances for specific account numbers.

    For asset accounts (1xxx): debit - credit (positive = balance exists).
    For liability accounts (2xxx): credit - debit (positive = balance exists).
    """
    total = Decimal('0')
    for acct_num in account_numbers:
        account = Account.query.filter_by(
            company_id=company_id, account_number=acct_num
        ).first()
        if not account:
            continue
        row = db.session.query(
            func.coalesce(func.sum(VerificationRow.debit), 0) -
            func.coalesce(func.sum(VerificationRow.credit), 0)
        ).join(Verification).filter(
            VerificationRow.account_id == account.id,
            Verification.fiscal_year_id == fiscal_year_id,
        ).scalar()
        balance = Decimal(str(row)) if row else Decimal('0')
        # For liability/equity accounts (2xxx), invert sign
        if acct_num.startswith('2'):
            balance = -balance
        if balance > 0:
            total += balance
    return total


def _sum_account_range(company_id, fiscal_year_id, prefix):
    """Sum absolute balances for accounts starting with prefix."""
    accounts = Account.query.filter(
        Account.company_id == company_id,
        Account.account_number.like(f'{prefix}%'),
    ).all()
    total = Decimal('0')
    for account in accounts:
        row = db.session.query(
            func.coalesce(func.sum(VerificationRow.debit), 0) -
            func.coalesce(func.sum(VerificationRow.credit), 0)
        ).join(Verification).filter(
            VerificationRow.account_id == account.id,
            Verification.fiscal_year_id == fiscal_year_id,
        ).scalar()
        balance = Decimal(str(row)) if row else Decimal('0')
        total += abs(balance)
    return total


def confirm_match(match_id, fiscal_year_id=None):
    """Confirm an intercompany match and auto-create elimination."""
    match = db.session.get(IntercompanyMatch, match_id)
    if not match or match.status != 'suggested':
        return None

    fy_id = fiscal_year_id or match.fiscal_year_id
    elimination = create_elimination(
        group_id=match.group_id,
        fiscal_year_id=fy_id,
        from_company_id=match.company_a_id,
        to_company_id=match.company_b_id,
        account_number='1660',
        amount=match.amount,
        description=match.description,
    )
    match.status = 'confirmed'
    match.elimination_id = elimination.id
    db.session.commit()
    return match


def reject_match(match_id):
    """Reject a suggested intercompany match."""
    match = db.session.get(IntercompanyMatch, match_id)
    if not match or match.status != 'suggested':
        return None
    match.status = 'rejected'
    db.session.commit()
    return match


def get_pending_matches(group_id, fiscal_year_id=None):
    """Get all suggested (pending) intercompany matches for a group."""
    query = IntercompanyMatch.query.filter_by(
        group_id=group_id, status='suggested'
    )
    if fiscal_year_id:
        query = query.filter_by(fiscal_year_id=fiscal_year_id)
    return query.order_by(IntercompanyMatch.amount.desc()).all()


# ---------------------------------------------------------------------------
# Goodwill
# ---------------------------------------------------------------------------

def register_acquisition(group_id, company_id, acquisition_date, purchase_price,
                         net_assets_at_acquisition, amortization_period_months=60):
    """Register an acquisition with goodwill calculation."""
    member = ConsolidationGroupMember.query.filter_by(
        group_id=group_id, company_id=company_id
    ).first()

    ownership_pct = float(member.ownership_pct) if member else 100.0
    net_assets_share = float(net_assets_at_acquisition) * ownership_pct / 100.0
    goodwill = float(purchase_price) - net_assets_share

    entry = AcquisitionGoodwill(
        group_id=group_id,
        company_id=company_id,
        acquisition_date=acquisition_date,
        purchase_price=Decimal(str(purchase_price)),
        net_assets_at_acquisition=Decimal(str(net_assets_at_acquisition)),
        goodwill_amount=Decimal(str(max(goodwill, 0))),
        amortization_period_months=amortization_period_months,
        accumulated_amortization=Decimal('0'),
    )
    db.session.add(entry)
    db.session.commit()
    return entry


def calculate_goodwill_amortization(goodwill_id, months=1):
    """Calculate and apply goodwill amortization for given months."""
    gw = db.session.get(AcquisitionGoodwill, goodwill_id)
    if not gw:
        return None

    monthly = gw.monthly_amortization
    amortization = Decimal(str(monthly * months))
    remaining = gw.remaining_goodwill

    if float(amortization) > remaining:
        amortization = Decimal(str(remaining))

    gw.accumulated_amortization = Decimal(str(float(gw.accumulated_amortization or 0))) + amortization
    db.session.commit()
    return float(amortization)


def _get_goodwill_amortization_for_year(group_id, fy_year):
    """Get total annual goodwill amortization for a group."""
    entries = AcquisitionGoodwill.query.filter_by(group_id=group_id).all()
    total = 0
    for gw in entries:
        if gw.acquisition_date.year <= fy_year:
            total += gw.monthly_amortization * 12
    return total


def get_total_remaining_goodwill(group_id):
    """Get total remaining goodwill for a group."""
    entries = AcquisitionGoodwill.query.filter_by(group_id=group_id).all()
    return sum(gw.remaining_goodwill for gw in entries)


def get_goodwill_entries(group_id):
    """Get all goodwill entries for a group."""
    return AcquisitionGoodwill.query.filter_by(group_id=group_id).order_by(
        AcquisitionGoodwill.acquisition_date
    ).all()


# ---------------------------------------------------------------------------
# Cash Flow Statement (indirect method)
# ---------------------------------------------------------------------------

def get_consolidated_cash_flow(group_id, fy_year):
    """Generate consolidated cash flow statement using indirect method.

    Operating: result + depreciation + working capital changes
    Investing: fixed asset changes
    Financing: equity + long-term debt changes
    """
    group = db.session.get(ConsolidationGroup, group_id)
    if not group:
        return None

    # We need current and prior year balance sheets
    pnl_data = get_consolidated_pnl(group_id, fy_year)
    bs_current = get_consolidated_balance_sheet(group_id, fy_year)
    bs_prior = get_consolidated_balance_sheet(group_id, fy_year - 1)

    if not pnl_data or not bs_current:
        return None

    # No members with data → no meaningful cash flow
    if not pnl_data.get('members'):
        return None

    result_before_tax = pnl_data.get('result_before_tax', 0)

    # Depreciation adjustments (non-cash: add back 78xx expense accounts)
    depreciation = 0
    for member in group.members:
        fy = _get_fy_for_company(member.company_id, fy_year)
        if not fy:
            continue
        depr_accounts = Account.query.filter(
            Account.company_id == member.company_id,
            Account.account_number.like('78%'),
        ).all()
        for acct in depr_accounts:
            row = db.session.query(
                func.coalesce(func.sum(VerificationRow.debit), 0) -
                func.coalesce(func.sum(VerificationRow.credit), 0)
            ).join(Verification).filter(
                VerificationRow.account_id == acct.id,
                Verification.fiscal_year_id == fy.id,
            ).scalar()
            depreciation += float(row) if row else 0

    # Working capital changes (if prior year exists)
    current_assets_change = 0
    current_liabilities_change = 0
    fixed_assets_change = 0
    equity_change = 0
    lt_debt_change = 0

    if bs_prior:
        # Current assets (excl cash 19xx)
        ca_current = bs_current['sections'].get('Omsättningstillgångar', {}).get('total', 0)
        ca_prior = bs_prior['sections'].get('Omsättningstillgångar', {}).get('total', 0)
        current_assets_change = ca_current - ca_prior

        # Current liabilities
        cl_current = bs_current['sections'].get('Kortfristiga skulder', {}).get('total', 0)
        cl_prior = bs_prior['sections'].get('Kortfristiga skulder', {}).get('total', 0)
        current_liabilities_change = cl_current - cl_prior

        # Fixed assets
        fa_current = bs_current['sections'].get('Anläggningstillgångar', {}).get('total', 0)
        fa_prior = bs_prior['sections'].get('Anläggningstillgångar', {}).get('total', 0)
        fixed_assets_change = fa_current - fa_prior

        # Equity
        eq_current = bs_current['sections'].get('Eget kapital', {}).get('total', 0)
        eq_prior = bs_prior['sections'].get('Eget kapital', {}).get('total', 0)
        equity_change = eq_current - eq_prior

        # Long-term debt
        ltd_current = bs_current['sections'].get('Långfristiga skulder', {}).get('total', 0)
        ltd_prior = bs_prior['sections'].get('Långfristiga skulder', {}).get('total', 0)
        lt_debt_change = ltd_current - ltd_prior

    # Operating activities
    working_capital_adj = -current_assets_change + current_liabilities_change
    operating = result_before_tax + depreciation + working_capital_adj

    # Investing activities (change in fixed assets + depreciation = net investment)
    investing = -(fixed_assets_change + depreciation)

    # Financing activities
    financing = equity_change - result_before_tax + lt_debt_change

    total_cash_flow = operating + investing + financing

    return {
        'group': group,
        'fy_year': fy_year,
        'result_before_tax': result_before_tax,
        'depreciation': depreciation,
        'working_capital_adj': working_capital_adj,
        'operating': operating,
        'investing': investing,
        'financing': financing,
        'total_cash_flow': total_cash_flow,
        'current_assets_change': current_assets_change,
        'current_liabilities_change': current_liabilities_change,
        'fixed_assets_change': fixed_assets_change,
        'equity_change': equity_change,
        'lt_debt_change': lt_debt_change,
    }


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def export_consolidated_report(group_id, fy_year, report_type):
    """Export consolidated report to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if report_type == 'pnl':
        data = get_consolidated_pnl(group_id, fy_year)
    elif report_type == 'cash_flow':
        data = get_consolidated_cash_flow(group_id, fy_year)
    else:
        data = get_consolidated_balance_sheet(group_id, fy_year)

    if not data:
        return None

    wb = Workbook()
    ws = wb.active
    bold = Font(bold=True)

    ws.append([f'Koncernrapport: {data["group"].name}'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'Räkenskapsår {fy_year}'])
    ws.append([])

    if report_type == 'pnl':
        ws.title = 'Koncern-RR'
        ws.append(['Avsnitt', 'Koncerntotal'])
        for cell in ws[ws.max_row]:
            cell.font = bold

        for section_name, section in data['sections'].items():
            ws.append([section_name, round(section['total'], 2)])

        ws.append([])
        ws.append(['Rörelseresultat', round(data['operating_result'], 2)])
        ws.append(['Resultat före skatt', round(data['result_before_tax'], 2)])
        ws.append(['Elimineringar', round(data['elimination_total'], 2)])
        ws.append(['Goodwill-avskrivning', round(data.get('goodwill_amortization', 0), 2)])
        ws.append(['Minoritetsintresse', round(data.get('minority_pnl', 0), 2)])
        ws.append(['Justerat resultat', round(data['adjusted_result'], 2)])
        ws[ws.max_row][0].font = bold

        ws.append([])
        ws.append(['Medlemsföretag'])
        ws[ws.max_row][0].font = bold
        for m in data['members']:
            ws.append([m['company'].name, f'{m["ownership_pct"]}%', m.get('consolidation_method', '')])

    elif report_type == 'cash_flow':
        ws.title = 'Kassaflöde'
        ws.append(['Post', 'Belopp'])
        for cell in ws[ws.max_row]:
            cell.font = bold

        ws.append(['Resultat före skatt', round(data['result_before_tax'], 2)])
        ws.append(['Avskrivningar', round(data['depreciation'], 2)])
        ws.append(['Rörelsekapitalförändring', round(data['working_capital_adj'], 2)])
        ws.append(['Kassaflöde från rörelsen', round(data['operating'], 2)])
        ws[ws.max_row][0].font = bold
        ws.append([])
        ws.append(['Kassaflöde från investeringar', round(data['investing'], 2)])
        ws[ws.max_row][0].font = bold
        ws.append([])
        ws.append(['Kassaflöde från finansiering', round(data['financing'], 2)])
        ws[ws.max_row][0].font = bold
        ws.append([])
        ws.append(['Totalt kassaflöde', round(data['total_cash_flow'], 2)])
        ws[ws.max_row][0].font = Font(bold=True, size=12)

    else:
        ws.title = 'Koncern-BR'
        ws.append(['Avsnitt', 'Koncerntotal'])
        for cell in ws[ws.max_row]:
            cell.font = bold

        for section_name, section in data['sections'].items():
            ws.append([section_name, round(section['total'], 2)])

        if data.get('remaining_goodwill', 0) > 0:
            ws.append(['  varav goodwill', round(data['remaining_goodwill'], 2)])

        ws.append([])
        ws.append(['Summa tillgångar', round(data['total_assets'], 2)])
        ws.append(['Summa EK + skulder', round(data['total_equity_liabilities'], 2)])
        if data.get('minority_equity', 0):
            ws.append(['Minoritetsintresse (EK)', round(data['minority_equity'], 2)])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
