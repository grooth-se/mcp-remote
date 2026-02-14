"""Deklaration (yearly tax return) service — INK2 for AB, INK4 for HB.

Auto-populates from accounting data, supports manual tax adjustments,
calculates taxable income and corporate tax.
"""

from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from sqlalchemy import func

from app.extensions import db
from app.models.tax import TaxReturn, TaxReturnAdjustment, TAX_RATES
from app.models.accounting import Account, Verification, VerificationRow, FiscalYear
from app.models.company import Company


# ---------------------------------------------------------------------------
# P&L extraction helpers
# ---------------------------------------------------------------------------

def _sum_accounts(company_id, fiscal_year_id, prefix_list, sign='debit'):
    """Sum account balances for given account number prefixes.

    sign='debit': returns debit - credit (expenses)
    sign='credit': returns credit - debit (revenue)
    """
    total = Decimal('0')
    for prefix in prefix_list:
        rows = db.session.query(
            func.coalesce(func.sum(VerificationRow.debit), 0).label('d'),
            func.coalesce(func.sum(VerificationRow.credit), 0).label('c'),
        ).join(
            Verification, Verification.id == VerificationRow.verification_id
        ).join(
            Account, Account.id == VerificationRow.account_id
        ).filter(
            Account.company_id == company_id,
            Verification.fiscal_year_id == fiscal_year_id,
            Account.account_number.like(f'{prefix}%'),
        ).first()

        if rows:
            if sign == 'credit':
                total += Decimal(str(rows.c)) - Decimal(str(rows.d))
            else:
                total += Decimal(str(rows.d)) - Decimal(str(rows.c))

    return total


def _extract_pnl_data(company_id, fiscal_year_id):
    """Extract P&L data from accounting, returning a dict of tax return fields."""
    # Revenue (3xxx) — credit side
    net_revenue = _sum_accounts(company_id, fiscal_year_id, ['3'], sign='credit')

    # Cost of goods sold (4xxx) — debit side
    cogs = _sum_accounts(company_id, fiscal_year_id, ['4'], sign='debit')

    # External costs (5xxx-6xxx) — debit side
    external = _sum_accounts(company_id, fiscal_year_id, ['5', '6'], sign='debit')

    # Personnel costs (7xxx excl depreciation 78xx) — debit side
    personnel = _sum_accounts(company_id, fiscal_year_id, ['70', '71', '72', '73', '74', '75', '76', '77'], sign='debit')

    # Depreciation (78xx) — debit side
    depreciation = _sum_accounts(company_id, fiscal_year_id, ['78'], sign='debit')

    # Financial income (80xx-83xx) — credit side
    fin_income = _sum_accounts(company_id, fiscal_year_id, ['80', '81', '82', '83'], sign='credit')

    # Financial expenses (84xx-89xx) — debit side
    fin_expense = _sum_accounts(company_id, fiscal_year_id, ['84', '85', '86', '87', '88', '89'], sign='debit')

    operating_expenses = cogs + external + personnel
    net_income = net_revenue - operating_expenses - depreciation + fin_income - fin_expense

    return {
        'net_revenue': net_revenue,
        'other_operating_income': Decimal('0'),  # User can adjust manually
        'operating_expenses': operating_expenses,
        'depreciation_booked': depreciation,
        'financial_income': fin_income,
        'financial_expenses': fin_expense,
        'net_income': net_income,
    }


# ---------------------------------------------------------------------------
# CRUD
# ---------------------------------------------------------------------------

def create_tax_return(company_id, fiscal_year_id, created_by=None):
    """Create a new tax return, auto-populated from accounting data."""
    company = db.session.get(Company, company_id)
    fy = db.session.get(FiscalYear, fiscal_year_id)
    if not company or not fy:
        return None

    # Determine return type
    if company.company_type == 'AB':
        return_type = 'ink2'
    elif company.company_type in ('HB', 'KB'):
        return_type = 'ink4'
    else:
        return_type = 'ink2'  # Default to INK2

    # Check for existing draft
    existing = TaxReturn.query.filter_by(
        company_id=company_id, fiscal_year_id=fiscal_year_id
    ).first()
    if existing:
        return existing

    # Extract P&L data
    pnl = _extract_pnl_data(company_id, fiscal_year_id)

    tax_rate = TAX_RATES.get(company.company_type, Decimal('0.206'))

    # Get previous year's deficit from prior tax return
    previous_deficit = _get_previous_deficit(company_id, fy.year)

    # Calculate taxable income
    taxable_before = pnl['net_income']  # Start with net income
    taxable_income = max(Decimal('0'), taxable_before - previous_deficit)
    calculated_tax = (taxable_income * tax_rate).quantize(Decimal('1'))

    tax_return = TaxReturn(
        company_id=company_id,
        fiscal_year_id=fiscal_year_id,
        return_type=return_type,
        tax_year=fy.year,
        status='draft',
        net_revenue=pnl['net_revenue'],
        other_operating_income=pnl['other_operating_income'],
        operating_expenses=pnl['operating_expenses'],
        depreciation_booked=pnl['depreciation_booked'],
        financial_income=pnl['financial_income'],
        financial_expenses=pnl['financial_expenses'],
        net_income=pnl['net_income'],
        non_deductible_expenses=Decimal('0'),
        non_taxable_income=Decimal('0'),
        depreciation_tax_diff=Decimal('0'),
        other_adjustments_add=Decimal('0'),
        other_adjustments_deduct=Decimal('0'),
        taxable_income_before_deficit=taxable_before,
        previous_deficit=previous_deficit,
        taxable_income=taxable_income,
        tax_rate=tax_rate,
        calculated_tax=calculated_tax,
        created_by=created_by,
    )
    db.session.add(tax_return)
    db.session.commit()
    return tax_return


def _get_previous_deficit(company_id, current_year):
    """Get accumulated deficit from previous year's tax return."""
    prev = TaxReturn.query.filter_by(company_id=company_id).filter(
        TaxReturn.tax_year < current_year
    ).order_by(TaxReturn.tax_year.desc()).first()

    if prev and prev.taxable_income < 0:
        return abs(prev.taxable_income)
    return Decimal('0')


def get_tax_return(return_id):
    """Get a single tax return with adjustments."""
    return db.session.get(TaxReturn, return_id)


def get_tax_returns(company_id):
    """List all tax returns for a company."""
    return TaxReturn.query.filter_by(company_id=company_id).order_by(
        TaxReturn.tax_year.desc()
    ).all()


def refresh_from_accounting(return_id):
    """Re-extract P&L data from accounting and recalculate."""
    tr = db.session.get(TaxReturn, return_id)
    if not tr or tr.status != 'draft':
        return None

    pnl = _extract_pnl_data(tr.company_id, tr.fiscal_year_id)
    tr.net_revenue = pnl['net_revenue']
    tr.other_operating_income = pnl['other_operating_income']
    tr.operating_expenses = pnl['operating_expenses']
    tr.depreciation_booked = pnl['depreciation_booked']
    tr.financial_income = pnl['financial_income']
    tr.financial_expenses = pnl['financial_expenses']
    tr.net_income = pnl['net_income']

    _recalculate(tr)
    db.session.commit()
    return tr


def update_adjustments(return_id, data):
    """Update manual adjustment fields and recalculate."""
    tr = db.session.get(TaxReturn, return_id)
    if not tr or tr.status != 'draft':
        return None

    fields = ['non_deductible_expenses', 'non_taxable_income',
              'depreciation_tax_diff', 'other_adjustments_add',
              'other_adjustments_deduct', 'previous_deficit', 'notes']
    for f in fields:
        if f in data and data[f] is not None:
            setattr(tr, f, data[f])

    _recalculate(tr)
    db.session.commit()
    return tr


def add_adjustment_line(return_id, adjustment_type, description, amount, sru_code=None):
    """Add a custom adjustment line item."""
    tr = db.session.get(TaxReturn, return_id)
    if not tr or tr.status != 'draft':
        return None

    adj = TaxReturnAdjustment(
        tax_return_id=return_id,
        adjustment_type=adjustment_type,
        description=description,
        amount=Decimal(str(amount)),
        sru_code=sru_code,
    )
    db.session.add(adj)

    # Update totals
    _sync_adjustment_totals(tr)
    _recalculate(tr)
    db.session.commit()
    return adj


def remove_adjustment_line(adjustment_id):
    """Remove a custom adjustment line item."""
    adj = db.session.get(TaxReturnAdjustment, adjustment_id)
    if not adj:
        return False

    tr = adj.tax_return
    if tr.status != 'draft':
        return False

    db.session.delete(adj)
    db.session.flush()

    _sync_adjustment_totals(tr)
    _recalculate(tr)
    db.session.commit()
    return True


def _sync_adjustment_totals(tr):
    """Sum adjustment line items into the tax return's add/deduct fields."""
    add_total = Decimal('0')
    deduct_total = Decimal('0')
    for adj in tr.adjustments:
        if adj.adjustment_type == 'add':
            add_total += adj.amount
        else:
            deduct_total += adj.amount
    tr.other_adjustments_add = add_total
    tr.other_adjustments_deduct = deduct_total


def _recalculate(tr):
    """Recalculate taxable income and tax from current fields."""
    # Taxable income = net_income + adjustments_add - adjustments_deduct
    # + non_deductible_expenses - non_taxable_income + depreciation_tax_diff
    taxable_before = (
        tr.net_income
        + tr.non_deductible_expenses
        - tr.non_taxable_income
        + tr.depreciation_tax_diff
        + tr.other_adjustments_add
        - tr.other_adjustments_deduct
    )
    tr.taxable_income_before_deficit = taxable_before

    # Apply previous year deficit (only reduces positive income)
    if taxable_before > 0 and tr.previous_deficit > 0:
        deficit_used = min(taxable_before, tr.previous_deficit)
        tr.taxable_income = taxable_before - deficit_used
    else:
        tr.taxable_income = taxable_before

    # Tax only on positive income
    if tr.taxable_income > 0:
        tr.calculated_tax = (tr.taxable_income * tr.tax_rate).quantize(Decimal('1'))
    else:
        tr.calculated_tax = Decimal('0')


def submit_tax_return(return_id):
    """Mark a tax return as submitted."""
    tr = db.session.get(TaxReturn, return_id)
    if not tr or tr.status != 'draft':
        return None
    tr.status = 'submitted'
    tr.submitted_at = datetime.now(timezone.utc)
    db.session.commit()
    return tr


def approve_tax_return(return_id):
    """Mark a tax return as approved (after Skatteverket confirmation)."""
    tr = db.session.get(TaxReturn, return_id)
    if not tr or tr.status != 'submitted':
        return None
    tr.status = 'approved'
    db.session.commit()
    return tr


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def export_tax_return_excel(return_id):
    """Export a tax return as formatted Excel."""
    tr = db.session.get(TaxReturn, return_id)
    if not tr:
        return None

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = f'Deklaration {tr.tax_year}'

    bold = Font(bold=True)
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    blue_fill = PatternFill(start_color='DBEEF4', end_color='DBEEF4', fill_type='solid')
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    # Header
    ws.append([f'Inkomstdeklaration {tr.return_type.upper()} — {tr.tax_year}'])
    ws['A1'].font = header_font
    ws.append([f'Företag: {tr.company.name}'])
    ws.append([f'Org.nr: {tr.company.org_number}'])
    ws.append([f'Status: {tr.status_label}'])
    ws.append([])

    # P&L section
    ws.append(['RESULTAT FRÅN BOKFÖRINGEN'])
    ws.cell(row=ws.max_row, column=1).font = section_font
    for cell in ws[ws.max_row]:
        cell.fill = blue_fill

    rows_pnl = [
        ('Nettoomsättning', tr.net_revenue),
        ('Övriga rörelseintäkter', tr.other_operating_income),
        ('Rörelsekostnader', tr.operating_expenses),
        ('Avskrivningar (bokförda)', tr.depreciation_booked),
        ('Finansiella intäkter', tr.financial_income),
        ('Finansiella kostnader', tr.financial_expenses),
    ]
    for label, val in rows_pnl:
        ws.append([label, '', float(val or 0)])

    ws.append(['Årets resultat', '', float(tr.net_income or 0)])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([])

    # Adjustments
    ws.append(['SKATTEMÄSSIGA JUSTERINGAR'])
    ws.cell(row=ws.max_row, column=1).font = section_font
    for cell in ws[ws.max_row]:
        cell.fill = blue_fill

    rows_adj = [
        ('Tillkommer: Ej avdragsgilla kostnader', tr.non_deductible_expenses),
        ('Avgår: Ej skattepliktiga intäkter', tr.non_taxable_income),
        ('Skillnad avskrivning (bokförd vs skattemässig)', tr.depreciation_tax_diff),
        ('Övriga justeringar (tillkommer)', tr.other_adjustments_add),
        ('Övriga justeringar (avgår)', tr.other_adjustments_deduct),
    ]
    for label, val in rows_adj:
        ws.append([label, '', float(val or 0)])

    # Adjustment line items
    if tr.adjustments:
        ws.append([])
        ws.append(['Detaljerade justeringar:'])
        ws.cell(row=ws.max_row, column=1).font = bold
        for adj in tr.adjustments:
            prefix = '+' if adj.adjustment_type == 'add' else '-'
            ws.append([f'  {prefix} {adj.description}', adj.sru_code or '', float(adj.amount)])

    ws.append([])

    # Taxable income
    ws.append(['BESKATTNINGSBART RESULTAT'])
    ws.cell(row=ws.max_row, column=1).font = section_font
    for cell in ws[ws.max_row]:
        cell.fill = green_fill

    ws.append(['Skattemässigt resultat före underskottsavdrag', '', float(tr.taxable_income_before_deficit or 0)])
    ws.append(['Underskott föregående år', '', float(tr.previous_deficit or 0)])
    ws.append(['Beskattningsbart resultat', '', float(tr.taxable_income or 0)])
    ws.cell(row=ws.max_row, column=1).font = bold

    ws.append([])
    ws.append([f'Skattesats: {float(tr.tax_rate) * 100:.1f}%'])
    ws.append(['Beräknad bolagsskatt', '', float(tr.calculated_tax or 0)])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

    # Notes
    if tr.notes:
        ws.append([])
        ws.append(['Anteckningar:'])
        ws.cell(row=ws.max_row, column=1).font = bold
        ws.append([tr.notes])

    # Auto-size columns
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def get_deklaration_summary(company_id, year):
    """Get a summary dict for dashboard display."""
    tr = TaxReturn.query.filter_by(company_id=company_id, tax_year=year).first()
    if not tr:
        return None
    return {
        'tax_year': tr.tax_year,
        'return_type': tr.return_type,
        'status': tr.status,
        'status_label': tr.status_label,
        'net_income': float(tr.net_income or 0),
        'taxable_income': float(tr.taxable_income or 0),
        'calculated_tax': float(tr.calculated_tax or 0),
    }
