"""Single-company cash flow statement (indirect method), monthly breakdown,
and simple cash flow projection.
"""

from io import BytesIO
from collections import OrderedDict
from datetime import date as date_type
from decimal import Decimal

from sqlalchemy import func, extract

from app.extensions import db
from app.models.accounting import Account, Verification, VerificationRow, FiscalYear
from app.services.report_service import (
    get_profit_and_loss, get_balance_sheet, _get_account_balances,
)


def get_cash_flow_statement(company_id, fiscal_year_id):
    """Generate a single-company cash flow statement using the indirect method.

    Compares current and prior year balance sheets to compute working capital
    changes, investing, and financing activities.

    Returns dict with operating, investing, financing sections + totals.
    """
    pnl = get_profit_and_loss(company_id, fiscal_year_id)
    bs = get_balance_sheet(company_id, fiscal_year_id)

    # Get prior year balance sheet
    fy = db.session.get(FiscalYear, fiscal_year_id)
    prior_fy = (FiscalYear.query
                .filter_by(company_id=company_id)
                .filter(FiscalYear.year < fy.year)
                .order_by(FiscalYear.year.desc())
                .first())

    prior_bs = None
    if prior_fy:
        prior_bs = get_balance_sheet(company_id, prior_fy.id)

    # Helper to get balance for account prefix from raw balances
    def _prefix_balance(balances, prefixes):
        return sum(b for a, b in balances if a.account_number[:2] in prefixes)

    raw_current = _get_account_balances(company_id, fiscal_year_id)
    raw_prior = _get_account_balances(company_id, prior_fy.id) if prior_fy else []

    # Current year values
    depreciation = sum(b for a, b in raw_current if a.account_number[:2] == '78')
    receivables = _prefix_balance(raw_current, ['15', '16'])
    inventory = _prefix_balance(raw_current, ['14'])
    cash_current = _prefix_balance(raw_current, ['19'])
    # Payables: credit balance (2440-2499) — need to negate debit-credit
    payables = -_prefix_balance(raw_current, ['24', '25', '26', '27', '28', '29'])

    # Prior year values
    receivables_prior = _prefix_balance(raw_prior, ['15', '16'])
    inventory_prior = _prefix_balance(raw_prior, ['14'])
    cash_prior = _prefix_balance(raw_prior, ['19'])
    payables_prior = -_prefix_balance(raw_prior, ['24', '25', '26', '27', '28', '29'])

    # Fixed assets (10-13xx) — current vs prior
    fixed_current = _prefix_balance(raw_current, ['10', '11', '12', '13'])
    fixed_prior = _prefix_balance(raw_prior, ['10', '11', '12', '13'])

    # Equity (20-21xx) and LT debt (22-23xx)
    equity_current = -_prefix_balance(raw_current, ['20', '21'])
    equity_prior = -_prefix_balance(raw_prior, ['20', '21'])
    lt_debt_current = -_prefix_balance(raw_current, ['22', '23'])
    lt_debt_prior = -_prefix_balance(raw_prior, ['22', '23'])

    result_before_tax = pnl['result_before_tax']

    # OPERATING ACTIVITIES (indirect method)
    adjustments = []

    if depreciation:
        adjustments.append({'label': 'Avskrivningar', 'amount': round(depreciation, 2)})

    delta_receivables = -(receivables - receivables_prior)
    if delta_receivables:
        adjustments.append({'label': 'Förändring kundfordringar', 'amount': round(delta_receivables, 2)})

    delta_inventory = -(inventory - inventory_prior)
    if delta_inventory:
        adjustments.append({'label': 'Förändring varulager', 'amount': round(delta_inventory, 2)})

    delta_payables = payables - payables_prior
    if delta_payables:
        adjustments.append({'label': 'Förändring leverantörsskulder', 'amount': round(delta_payables, 2)})

    operating_total = result_before_tax + sum(a['amount'] for a in adjustments)

    # INVESTING ACTIVITIES
    # Net fixed asset change + depreciation = investment
    net_investment = -(fixed_current - fixed_prior + depreciation)
    investing_items = []
    if net_investment:
        investing_items.append({
            'label': 'Förvärv/försäljning av anläggningstillgångar',
            'amount': round(net_investment, 2),
        })
    investing_total = sum(i['amount'] for i in investing_items)

    # FINANCING ACTIVITIES
    financing_items = []
    delta_equity = equity_current - equity_prior - result_before_tax
    if abs(delta_equity) > 0.01:
        financing_items.append({
            'label': 'Förändring eget kapital (exkl årets resultat)',
            'amount': round(delta_equity, 2),
        })

    delta_lt_debt = lt_debt_current - lt_debt_prior
    if abs(delta_lt_debt) > 0.01:
        financing_items.append({
            'label': 'Förändring långfristiga skulder',
            'amount': round(delta_lt_debt, 2),
        })
    financing_total = sum(i['amount'] for i in financing_items)

    total_cash_flow = round(operating_total + investing_total + financing_total, 2)

    return {
        'operating': {
            'result_before_tax': round(result_before_tax, 2),
            'adjustments': adjustments,
            'total': round(operating_total, 2),
        },
        'investing': {
            'line_items': investing_items,
            'total': round(investing_total, 2),
        },
        'financing': {
            'line_items': financing_items,
            'total': round(financing_total, 2),
        },
        'total_cash_flow': total_cash_flow,
        'opening_cash': round(cash_prior, 2),
        'closing_cash': round(cash_current, 2),
    }


def get_monthly_cash_flow(company_id, fiscal_year_id):
    """Monthly cash flow breakdown from bank/cash account (19xx) movements.

    Classifies each verification's counterpart accounts to determine
    operating / investing / financing.

    Returns dict with labels and monthly arrays.
    """
    fy = db.session.get(FiscalYear, fiscal_year_id)
    months = list(range(1, 13))
    labels = ['Jan', 'Feb', 'Mar', 'Apr', 'Maj', 'Jun',
              'Jul', 'Aug', 'Sep', 'Okt', 'Nov', 'Dec']

    operating = [0.0] * 12
    investing = [0.0] * 12
    financing = [0.0] * 12

    # Get all verifications with cash account rows
    cash_rows = (db.session.query(
        Verification.id,
        Verification.verification_date,
        VerificationRow.debit,
        VerificationRow.credit,
        Account.account_number,
    ).join(VerificationRow, VerificationRow.verification_id == Verification.id)
     .join(Account, Account.id == VerificationRow.account_id)
     .filter(
         Verification.company_id == company_id,
         Verification.fiscal_year_id == fiscal_year_id,
         Account.account_number.like('19%'),
     ).all())

    # Group by verification ID
    cash_by_ver = {}
    for ver_id, ver_date, debit, credit, acct_num in cash_rows:
        if ver_id not in cash_by_ver:
            cash_by_ver[ver_id] = {'date': ver_date, 'amount': 0.0}
        cash_by_ver[ver_id]['amount'] += float(debit) - float(credit)

    # Get non-cash rows for classification
    for ver_id, ver_info in cash_by_ver.items():
        month_idx = ver_info['date'].month - 1
        amount = ver_info['amount']

        # Get counterpart accounts
        counterparts = (db.session.query(Account.account_number)
                        .join(VerificationRow, VerificationRow.account_id == Account.id)
                        .filter(
                            VerificationRow.verification_id == ver_id,
                            ~Account.account_number.like('19%'),
                        ).all())

        category = _classify_counterparts([c[0] for c in counterparts])

        if category == 'investing':
            investing[month_idx] += amount
        elif category == 'financing':
            financing[month_idx] += amount
        else:
            operating[month_idx] += amount

    net = [round(o + i + f, 2) for o, i, f in zip(operating, investing, financing)]
    cumulative = []
    running = 0.0
    for n in net:
        running += n
        cumulative.append(round(running, 2))

    return {
        'labels': labels,
        'operating': [round(v, 2) for v in operating],
        'investing': [round(v, 2) for v in investing],
        'financing': [round(v, 2) for v in financing],
        'net': net,
        'cumulative': cumulative,
    }


def get_cash_flow_forecast(company_id, fiscal_year_id, forecast_months=3):
    """Simple cash flow projection based on rolling average.

    Returns actual monthly data + forecast for additional months.
    """
    monthly = get_monthly_cash_flow(company_id, fiscal_year_id)

    # Find the last month with data
    actual_months = [i for i, v in enumerate(monthly['net']) if v != 0]
    if not actual_months:
        return {
            'labels': monthly['labels'],
            'actual': monthly['net'],
            'forecast': [None] * 12,
            'avg_monthly_cf': 0.0,
        }

    last_month = max(actual_months)
    # Average of months with data
    data_values = [monthly['net'][i] for i in actual_months]
    avg_cf = sum(data_values) / len(data_values) if data_values else 0.0

    # Build forecast: None for actual months, avg for future
    forecast = [None] * 12
    for i in range(last_month + 1, min(last_month + 1 + forecast_months, 12)):
        forecast[i] = round(avg_cf, 2)

    return {
        'labels': monthly['labels'],
        'actual': monthly['net'],
        'forecast': forecast,
        'avg_monthly_cf': round(avg_cf, 2),
    }


def export_cashflow_to_excel(cf_data, company_name, fiscal_year):
    """Export cash flow statement to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = 'Kassaflödesanalys'
    bold = Font(bold=True)
    header_font = Font(bold=True, size=14)

    ws.append([company_name])
    ws['A1'].font = header_font
    ws.append([f'Kassaflödesanalys {fiscal_year.start_date} - {fiscal_year.end_date}'])
    ws.append([])

    # Operating
    ws.append(['DEN LÖPANDE VERKSAMHETEN'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append(['Resultat före skatt', '', cf_data['operating']['result_before_tax']])
    for adj in cf_data['operating']['adjustments']:
        ws.append([adj['label'], '', adj['amount']])
    ws.append(['Kassaflöde från den löpande verksamheten', '', cf_data['operating']['total']])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([])

    # Investing
    ws.append(['INVESTERINGSVERKSAMHETEN'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    for item in cf_data['investing']['line_items']:
        ws.append([item['label'], '', item['amount']])
    ws.append(['Kassaflöde från investeringsverksamheten', '', cf_data['investing']['total']])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([])

    # Financing
    ws.append(['FINANSIERINGSVERKSAMHETEN'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    for item in cf_data['financing']['line_items']:
        ws.append([item['label'], '', item['amount']])
    ws.append(['Kassaflöde från finansieringsverksamheten', '', cf_data['financing']['total']])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([])

    # Summary
    ws.append(['ÅRETS KASSAFLÖDE', '', cf_data['total_cash_flow']])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.append([])
    ws.append(['Likvida medel vid årets början', '', cf_data['opening_cash']])
    ws.append(['Likvida medel vid årets slut', '', cf_data['closing_cash']])

    for col in ws.columns:
        max_length = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def get_enhanced_cash_flow_forecast(company_id, fiscal_year_id, forecast_months=3):
    """Enhanced cash flow forecast with seasonal patterns, known obligations,
    and confidence intervals.

    Returns dict with actual, forecast, confidence_upper, confidence_lower,
    known_obligations, labels.
    """
    import statistics

    fy = db.session.get(FiscalYear, fiscal_year_id)
    monthly = get_monthly_cash_flow(company_id, fiscal_year_id)

    # Find actual months with data
    actual_months = [i for i, v in enumerate(monthly['net']) if v != 0]

    if not actual_months:
        return {
            'labels': monthly['labels'],
            'actual': monthly['net'],
            'forecast': [None] * 12,
            'confidence_upper': [None] * 12,
            'confidence_lower': [None] * 12,
            'known_obligations': [0.0] * 12,
            'avg_monthly_cf': 0.0,
            'has_seasonal': False,
        }

    last_month = max(actual_months)
    data_values = [monthly['net'][i] for i in actual_months]
    avg_cf = sum(data_values) / len(data_values)

    # Compute stddev for confidence intervals
    if len(data_values) >= 2:
        try:
            stddev = statistics.stdev(data_values)
        except statistics.StatisticsError:
            stddev = 0.0
    else:
        stddev = 0.0

    # Seasonal patterns from prior year
    seasonal_factors = {}
    has_seasonal = False
    prior_fy = (FiscalYear.query
                .filter_by(company_id=company_id)
                .filter(FiscalYear.year < fy.year)
                .order_by(FiscalYear.year.desc())
                .first())

    if prior_fy:
        prior_monthly = get_monthly_cash_flow(company_id, prior_fy.id)
        prior_values = prior_monthly['net']
        prior_data = [i for i, v in enumerate(prior_values) if v != 0]
        if prior_data:
            prior_avg = sum(prior_values[i] for i in prior_data) / len(prior_data)
            if abs(prior_avg) > 0.01:
                for m in range(12):
                    seasonal_factors[m] = prior_values[m] / prior_avg if prior_values[m] != 0 else 1.0
                has_seasonal = True

    # Known obligations: pending supplier invoices + recurring templates + tax payments
    known_obligations = [0.0] * 12
    _add_known_obligations(company_id, fy, known_obligations)

    # Build forecast
    forecast = [None] * 12
    confidence_upper = [None] * 12
    confidence_lower = [None] * 12

    for i in range(last_month + 1, min(last_month + 1 + forecast_months, 12)):
        base = avg_cf

        # Apply seasonal factor if available
        if has_seasonal and i in seasonal_factors:
            base = avg_cf * seasonal_factors[i]

        # Add known obligations
        base += known_obligations[i]

        forecast[i] = round(base, 2)
        confidence_upper[i] = round(base + stddev, 2)
        confidence_lower[i] = round(base - stddev, 2)

    return {
        'labels': monthly['labels'],
        'actual': monthly['net'],
        'forecast': forecast,
        'confidence_upper': confidence_upper,
        'confidence_lower': confidence_lower,
        'known_obligations': [round(v, 2) for v in known_obligations],
        'avg_monthly_cf': round(avg_cf, 2),
        'has_seasonal': has_seasonal,
    }


def _add_known_obligations(company_id, fy, obligations):
    """Add known future payments to monthly obligations array."""
    from app.models.invoice import SupplierInvoice
    from app.models.recurring_invoice import RecurringInvoiceTemplate
    from app.models.tax import TaxPayment

    # Pending supplier invoices due within FY
    pending = SupplierInvoice.query.filter_by(
        company_id=company_id, status='pending'
    ).filter(
        SupplierInvoice.due_date >= fy.start_date,
        SupplierInvoice.due_date <= fy.end_date,
    ).all()

    for inv in pending:
        if inv.due_date and inv.total_amount:
            month_idx = inv.due_date.month - 1
            obligations[month_idx] -= float(inv.total_amount)

    # Recurring invoice templates (expected income)
    templates = RecurringInvoiceTemplate.query.filter_by(
        company_id=company_id, active=True
    ).all()

    for tmpl in templates:
        if tmpl.next_date and fy.start_date <= tmpl.next_date <= fy.end_date:
            total = sum(float(li.unit_price or 0) * float(li.quantity or 1) for li in tmpl.line_items)
            if total > 0:
                month_idx = tmpl.next_date.month - 1
                obligations[month_idx] += total

    # Historical tax payments for recurring months (use prior year pattern)
    prior_payments = TaxPayment.query.filter_by(company_id=company_id).filter(
        db.extract('year', TaxPayment.payment_date) == fy.year - 1,
    ).all()

    for tp in prior_payments:
        if tp.payment_date and tp.amount:
            month_idx = tp.payment_date.month - 1
            obligations[month_idx] -= float(tp.amount)


def _classify_counterparts(account_numbers):
    """Classify a verification by its non-cash counterpart accounts.

    Returns 'operating', 'investing', or 'financing'.
    """
    for num in account_numbers:
        prefix = num[:2]
        if prefix in ('10', '11', '12', '13'):
            return 'investing'
        if prefix in ('20', '21', '22', '23'):
            return 'financing'
    return 'operating'
