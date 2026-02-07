"""Budget & Forecast service: grid editor, variance analysis, forecast."""

from decimal import Decimal
from io import BytesIO
from collections import OrderedDict
from sqlalchemy import func

from app.extensions import db
from app.models.budget import BudgetLine
from app.models.accounting import Account, Verification, VerificationRow, FiscalYear
from app.models.audit import AuditLog


def get_budget_grid(company_id, fiscal_year_id):
    """Get budget data as a grid: {account_id: {number, name, months: {1..12}, total}}."""
    # Get all P&L accounts (3xxx-8xxx) for this company
    accounts = Account.query.filter(
        Account.company_id == company_id,
        Account.active == True,
        Account.account_number >= '3000',
        Account.account_number <= '8999',
    ).order_by(Account.account_number).all()

    # Get existing budget lines
    lines = BudgetLine.query.filter_by(
        company_id=company_id,
        fiscal_year_id=fiscal_year_id,
    ).all()

    # Index by (account_id, month)
    line_map = {}
    for line in lines:
        line_map[(line.account_id, line.period_month)] = float(line.amount or 0)

    grid = OrderedDict()
    for acc in accounts:
        months = {}
        total = 0
        for m in range(1, 13):
            val = line_map.get((acc.id, m), 0)
            months[m] = val
            total += val

        # Only include accounts that have budget entries or are commonly used
        if total != 0 or acc.account_number[0] in ('3', '4', '5', '6', '7'):
            grid[acc.id] = {
                'account_id': acc.id,
                'number': acc.account_number,
                'name': acc.name,
                'months': months,
                'total': total,
            }

    return grid


def save_budget_grid(company_id, fiscal_year_id, grid_data, user_id):
    """Save budget grid data. grid_data = {account_id: {month: amount}}."""
    count = 0
    for account_id_str, months in grid_data.items():
        account_id = int(account_id_str)
        for month_str, amount_str in months.items():
            month = int(month_str)
            try:
                amount = Decimal(str(amount_str or 0))
            except Exception:
                amount = Decimal('0')

            existing = BudgetLine.query.filter_by(
                company_id=company_id,
                fiscal_year_id=fiscal_year_id,
                account_id=account_id,
                period_month=month,
            ).first()

            if existing:
                if float(existing.amount or 0) != float(amount):
                    existing.amount = amount
                    count += 1
            elif amount != 0:
                line = BudgetLine(
                    company_id=company_id,
                    fiscal_year_id=fiscal_year_id,
                    account_id=account_id,
                    period_month=month,
                    amount=amount,
                )
                db.session.add(line)
                count += 1

    if count > 0:
        audit = AuditLog(
            company_id=company_id, user_id=user_id,
            action='update', entity_type='budget',
            entity_id=fiscal_year_id,
            new_values={'lines_updated': count},
        )
        db.session.add(audit)
        db.session.commit()

    return count


def get_variance_analysis(company_id, fiscal_year_id):
    """Compare budget vs actual per account per month."""
    grid = get_budget_grid(company_id, fiscal_year_id)

    # Get actual data per account per month
    actuals = db.session.query(
        VerificationRow.account_id,
        func.extract('month', Verification.verification_date).label('month'),
        func.sum(VerificationRow.debit).label('total_debit'),
        func.sum(VerificationRow.credit).label('total_credit'),
    ).join(
        Verification, Verification.id == VerificationRow.verification_id
    ).filter(
        Verification.company_id == company_id,
        Verification.fiscal_year_id == fiscal_year_id,
    ).group_by(
        VerificationRow.account_id,
        func.extract('month', Verification.verification_date),
    ).all()

    actual_map = {}
    for row in actuals:
        actual_map[(row.account_id, int(row.month))] = float(row.total_debit - row.total_credit)

    results = []
    for account_id, data in grid.items():
        months = []
        budget_total = 0
        actual_total = 0
        for m in range(1, 13):
            budget = data['months'].get(m, 0)
            # For revenue (3xxx), actual = credit - debit (negative balance = revenue)
            raw_actual = actual_map.get((account_id, m), 0)
            if data['number'][0] == '3':
                actual = -raw_actual  # revenue shown as positive
            else:
                actual = raw_actual

            variance = actual - budget
            pct = round((variance / budget * 100), 1) if budget else 0

            months.append({
                'budget': budget,
                'actual': actual,
                'variance': variance,
                'pct': pct,
            })
            budget_total += budget
            actual_total += actual

        total_variance = actual_total - budget_total
        total_pct = round((total_variance / budget_total * 100), 1) if budget_total else 0

        results.append({
            'account_id': account_id,
            'number': data['number'],
            'name': data['name'],
            'months': months,
            'budget_total': budget_total,
            'actual_total': actual_total,
            'variance_total': total_variance,
            'variance_pct': total_pct,
        })

    return results


def get_forecast(company_id, fiscal_year_id):
    """Actual months + projected remaining months based on trend."""
    from datetime import date

    fy = db.session.get(FiscalYear, fiscal_year_id)
    if not fy:
        return None

    today = date.today()
    current_month = today.month

    grid = get_budget_grid(company_id, fiscal_year_id)

    # Get actual monthly totals (revenue - expenses)
    monthly_actual = {}
    for m in range(1, 13):
        rev = db.session.query(
            func.coalesce(func.sum(VerificationRow.credit - VerificationRow.debit), 0)
        ).join(
            Verification, Verification.id == VerificationRow.verification_id
        ).join(
            Account, Account.id == VerificationRow.account_id
        ).filter(
            Verification.company_id == company_id,
            Verification.fiscal_year_id == fiscal_year_id,
            func.extract('month', Verification.verification_date) == m,
            Account.account_number.like('3%'),
        ).scalar()

        exp = db.session.query(
            func.coalesce(func.sum(VerificationRow.debit - VerificationRow.credit), 0)
        ).join(
            Verification, Verification.id == VerificationRow.verification_id
        ).join(
            Account, Account.id == VerificationRow.account_id
        ).filter(
            Verification.company_id == company_id,
            Verification.fiscal_year_id == fiscal_year_id,
            func.extract('month', Verification.verification_date) == m,
            Account.account_number.like('4%') |
            Account.account_number.like('5%') |
            Account.account_number.like('6%') |
            Account.account_number.like('7%'),
        ).scalar()

        monthly_actual[m] = {
            'revenue': float(rev or 0),
            'expenses': float(exp or 0),
            'result': float((rev or 0) - (exp or 0)),
        }

    # Calculate average for projection
    actual_months = [m for m in range(1, current_month + 1) if monthly_actual[m]['revenue'] > 0 or monthly_actual[m]['expenses'] > 0]
    if actual_months:
        avg_revenue = sum(monthly_actual[m]['revenue'] for m in actual_months) / len(actual_months)
        avg_expenses = sum(monthly_actual[m]['expenses'] for m in actual_months) / len(actual_months)
    else:
        avg_revenue = 0
        avg_expenses = 0

    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'Maj', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Okt', 'Nov', 'Dec']

    labels = month_names
    actual_data = []
    forecast_data = []

    for m in range(1, 13):
        if m <= current_month:
            actual_data.append(monthly_actual[m]['result'])
            forecast_data.append(None)
        else:
            actual_data.append(None)
            forecast_data.append(round(avg_revenue - avg_expenses, 2))

    return {
        'labels': labels,
        'actual': actual_data,
        'forecast': forecast_data,
        'current_month': current_month,
        'avg_revenue': avg_revenue,
        'avg_expenses': avg_expenses,
    }


def copy_budget_from_year(company_id, source_fy_id, target_fy_id, user_id):
    """Copy budget lines from one fiscal year to another."""
    lines = BudgetLine.query.filter_by(
        company_id=company_id,
        fiscal_year_id=source_fy_id,
    ).all()

    count = 0
    for line in lines:
        existing = BudgetLine.query.filter_by(
            company_id=company_id,
            fiscal_year_id=target_fy_id,
            account_id=line.account_id,
            period_month=line.period_month,
        ).first()

        if not existing:
            new_line = BudgetLine(
                company_id=company_id,
                fiscal_year_id=target_fy_id,
                account_id=line.account_id,
                period_month=line.period_month,
                amount=line.amount,
                notes=line.notes,
            )
            db.session.add(new_line)
            count += 1

    if count > 0:
        audit = AuditLog(
            company_id=company_id, user_id=user_id,
            action='create', entity_type='budget',
            entity_id=target_fy_id,
            new_values={'copied_lines': count, 'source_fy': source_fy_id},
        )
        db.session.add(audit)
        db.session.commit()

    return count


def export_budget_to_excel(company_id, fiscal_year_id, company_name):
    """Export budget grid to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    grid = get_budget_grid(company_id, fiscal_year_id)
    fy = db.session.get(FiscalYear, fiscal_year_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Budget'

    bold = Font(bold=True)
    ws.append([company_name])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'Budget {fy.year}' if fy else 'Budget'])
    ws.append([])

    months = ['Konto', 'Kontonamn', 'Jan', 'Feb', 'Mar', 'Apr', 'Maj', 'Jun',
              'Jul', 'Aug', 'Sep', 'Okt', 'Nov', 'Dec', 'Totalt']
    ws.append(months)
    for cell in ws[ws.max_row]:
        cell.font = bold

    for acc_id, data in grid.items():
        row = [data['number'], data['name']]
        for m in range(1, 13):
            row.append(round(data['months'].get(m, 0), 2))
        row.append(round(data['total'], 2))
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_variance_to_excel(company_id, fiscal_year_id, company_name):
    """Export variance analysis to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    variance = get_variance_analysis(company_id, fiscal_year_id)
    fy = db.session.get(FiscalYear, fiscal_year_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Avvikelseanalys'

    bold = Font(bold=True)
    ws.append([company_name])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'Avvikelseanalys {fy.year}' if fy else 'Avvikelseanalys'])
    ws.append([])

    ws.append(['Konto', 'Kontonamn', 'Budget totalt', 'Utfall totalt', 'Avvikelse', 'Avvikelse %'])
    for cell in ws[ws.max_row]:
        cell.font = bold

    for item in variance:
        ws.append([
            item['number'], item['name'],
            round(item['budget_total'], 2),
            round(item['actual_total'], 2),
            round(item['variance_total'], 2),
            f"{item['variance_pct']}%",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
