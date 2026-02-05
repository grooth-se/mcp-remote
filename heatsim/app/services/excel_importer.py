"""Excel import/export service for material data."""
import io
import json
from typing import BinaryIO, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.extensions import db
from app.models import (
    SteelGrade, MaterialProperty, PhaseDiagram,
    PROPERTY_TYPE_CONSTANT, PROPERTY_TYPE_CURVE,
    DATA_SOURCE_STANDARD, DATA_SOURCE_SUBSEATEC
)


class ExcelImporter:
    """Import and export material data from/to Excel files."""

    # Standard property names for import/export
    STANDARD_PROPERTIES = [
        'thermal_conductivity',
        'specific_heat',
        'density',
        'emissivity',
        'youngs_modulus',
        'poissons_ratio',
        'thermal_expansion',
    ]

    # Units for standard properties
    PROPERTY_UNITS = {
        'thermal_conductivity': 'W/(m·K)',
        'specific_heat': 'J/(kg·K)',
        'density': 'kg/m³',
        'emissivity': '-',
        'youngs_modulus': 'GPa',
        'poissons_ratio': '-',
        'thermal_expansion': '1/K',
    }

    @classmethod
    def import_from_excel(cls, file: BinaryIO, data_source: str = DATA_SOURCE_STANDARD) -> dict:
        """Import material data from Excel file.

        Expected format:
        - Sheet "Steel Grades": designation, description
        - Sheet "Properties": designation, property_name, property_type, units, dependencies, data
        - Sheet "Phase Diagrams": designation, diagram_type, Ac1, Ac3, Ms, Mf

        Parameters
        ----------
        file : BinaryIO
            Excel file object
        data_source : str
            Data source for imported grades

        Returns
        -------
        dict
            Import results with counts and errors
        """
        results = {
            'grades_created': 0,
            'grades_updated': 0,
            'properties_created': 0,
            'diagrams_created': 0,
            'errors': []
        }

        try:
            xl = pd.ExcelFile(file)
        except Exception as e:
            results['errors'].append(f"Failed to read Excel file: {str(e)}")
            return results

        # Import steel grades
        if 'Steel Grades' in xl.sheet_names:
            cls._import_grades(xl, data_source, results)

        # Import properties
        if 'Properties' in xl.sheet_names:
            cls._import_properties(xl, data_source, results)

        # Import phase diagrams
        if 'Phase Diagrams' in xl.sheet_names:
            cls._import_diagrams(xl, data_source, results)

        return results

    @classmethod
    def _import_grades(cls, xl: pd.ExcelFile, data_source: str, results: dict):
        """Import steel grades from Excel."""
        df = xl.parse('Steel Grades')

        for _, row in df.iterrows():
            designation = str(row.get('designation', '')).strip()
            if not designation:
                continue

            # Check if grade exists
            existing = SteelGrade.query.filter_by(
                designation=designation,
                data_source=data_source
            ).first()

            if existing:
                existing.description = row.get('description', '')
                results['grades_updated'] += 1
            else:
                grade = SteelGrade(
                    designation=designation,
                    data_source=data_source,
                    description=row.get('description', '')
                )
                db.session.add(grade)
                results['grades_created'] += 1

        db.session.commit()

    @classmethod
    def _import_properties(cls, xl: pd.ExcelFile, data_source: str, results: dict):
        """Import material properties from Excel."""
        df = xl.parse('Properties')

        for _, row in df.iterrows():
            designation = str(row.get('designation', '')).strip()
            property_name = str(row.get('property_name', '')).strip()

            if not designation or not property_name:
                continue

            # Find the steel grade
            grade = SteelGrade.query.filter_by(
                designation=designation,
                data_source=data_source
            ).first()

            if not grade:
                results['errors'].append(f"Grade not found: {designation}")
                continue

            # Parse data
            data_str = row.get('data', '{}')
            if isinstance(data_str, str):
                try:
                    data = json.loads(data_str)
                except json.JSONDecodeError:
                    results['errors'].append(f"Invalid JSON data for {designation}/{property_name}")
                    continue
            else:
                data = {'value': float(data_str)} if pd.notna(data_str) else {}

            # Create or update property
            existing = MaterialProperty.query.filter_by(
                steel_grade_id=grade.id,
                property_name=property_name
            ).first()

            if existing:
                existing.property_type = row.get('property_type', PROPERTY_TYPE_CONSTANT)
                existing.units = row.get('units', '')
                existing.dependencies = row.get('dependencies', '')
                existing.set_data(data)
            else:
                prop = MaterialProperty(
                    steel_grade_id=grade.id,
                    property_name=property_name,
                    property_type=row.get('property_type', PROPERTY_TYPE_CONSTANT),
                    units=row.get('units', ''),
                    dependencies=row.get('dependencies', ''),
                )
                prop.set_data(data)
                db.session.add(prop)
                results['properties_created'] += 1

        db.session.commit()

    @classmethod
    def _import_diagrams(cls, xl: pd.ExcelFile, data_source: str, results: dict):
        """Import phase diagrams from Excel."""
        df = xl.parse('Phase Diagrams')

        for _, row in df.iterrows():
            designation = str(row.get('designation', '')).strip()
            diagram_type = str(row.get('diagram_type', 'CCT')).strip().upper()

            if not designation:
                continue

            # Find the steel grade
            grade = SteelGrade.query.filter_by(
                designation=designation,
                data_source=data_source
            ).first()

            if not grade:
                results['errors'].append(f"Grade not found for diagram: {designation}")
                continue

            # Parse transformation temperatures
            temps = {}
            for temp_name in ['Ac1', 'Ac3', 'Ms', 'Mf', 'Bs', 'Bf']:
                value = row.get(temp_name)
                if pd.notna(value):
                    temps[temp_name] = float(value)

            # Create diagram
            diagram = PhaseDiagram(
                steel_grade_id=grade.id,
                diagram_type=diagram_type
            )
            diagram.set_temps(temps)
            db.session.add(diagram)
            results['diagrams_created'] += 1

        db.session.commit()

    @classmethod
    def export_to_excel(cls, grade_ids: Optional[list] = None) -> bytes:
        """Export material data to Excel file.

        Parameters
        ----------
        grade_ids : list, optional
            List of grade IDs to export, or None for all

        Returns
        -------
        bytes
            Excel file contents
        """
        wb = Workbook()

        # Styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Query grades
        if grade_ids:
            grades = SteelGrade.query.filter(SteelGrade.id.in_(grade_ids)).all()
        else:
            grades = SteelGrade.query.all()

        # Sheet 1: Steel Grades
        ws_grades = wb.active
        ws_grades.title = 'Steel Grades'
        headers = ['designation', 'data_source', 'description']
        for col, header in enumerate(headers, 1):
            cell = ws_grades.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for row, grade in enumerate(grades, 2):
            ws_grades.cell(row=row, column=1, value=grade.designation)
            ws_grades.cell(row=row, column=2, value=grade.data_source)
            ws_grades.cell(row=row, column=3, value=grade.description or '')

        # Sheet 2: Properties
        ws_props = wb.create_sheet('Properties')
        headers = ['designation', 'property_name', 'property_type', 'units', 'dependencies', 'data']
        for col, header in enumerate(headers, 1):
            cell = ws_props.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row = 2
        for grade in grades:
            for prop in grade.properties:
                ws_props.cell(row=row, column=1, value=grade.designation)
                ws_props.cell(row=row, column=2, value=prop.property_name)
                ws_props.cell(row=row, column=3, value=prop.property_type)
                ws_props.cell(row=row, column=4, value=prop.units or '')
                ws_props.cell(row=row, column=5, value=prop.dependencies or '')
                ws_props.cell(row=row, column=6, value=prop.data)
                row += 1

        # Sheet 3: Phase Diagrams
        ws_diagrams = wb.create_sheet('Phase Diagrams')
        headers = ['designation', 'diagram_type', 'Ac1', 'Ac3', 'Ms', 'Mf', 'Bs', 'Bf']
        for col, header in enumerate(headers, 1):
            cell = ws_diagrams.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row = 2
        for grade in grades:
            for diagram in grade.phase_diagrams:
                temps = diagram.temps_dict
                ws_diagrams.cell(row=row, column=1, value=grade.designation)
                ws_diagrams.cell(row=row, column=2, value=diagram.diagram_type)
                ws_diagrams.cell(row=row, column=3, value=temps.get('Ac1'))
                ws_diagrams.cell(row=row, column=4, value=temps.get('Ac3'))
                ws_diagrams.cell(row=row, column=5, value=temps.get('Ms'))
                ws_diagrams.cell(row=row, column=6, value=temps.get('Mf'))
                ws_diagrams.cell(row=row, column=7, value=temps.get('Bs'))
                ws_diagrams.cell(row=row, column=8, value=temps.get('Bf'))
                row += 1

        # Auto-adjust column widths
        for ws in [ws_grades, ws_props, ws_diagrams]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @classmethod
    def get_template(cls) -> bytes:
        """Generate empty Excel template for data entry.

        Returns
        -------
        bytes
            Excel template file contents
        """
        wb = Workbook()

        # Styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(bold=True, color='FFFFFF')

        # Sheet 1: Steel Grades
        ws_grades = wb.active
        ws_grades.title = 'Steel Grades'
        headers = ['designation', 'description']
        for col, header in enumerate(headers, 1):
            cell = ws_grades.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill

        # Example row
        ws_grades.cell(row=2, column=1, value='AISI 4340')
        ws_grades.cell(row=2, column=2, value='High-strength low-alloy steel')

        # Sheet 2: Properties
        ws_props = wb.create_sheet('Properties')
        headers = ['designation', 'property_name', 'property_type', 'units', 'dependencies', 'data']
        for col, header in enumerate(headers, 1):
            cell = ws_props.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill

        # Example rows
        ws_props.cell(row=2, column=1, value='AISI 4340')
        ws_props.cell(row=2, column=2, value='density')
        ws_props.cell(row=2, column=3, value='constant')
        ws_props.cell(row=2, column=4, value='kg/m³')
        ws_props.cell(row=2, column=5, value='')
        ws_props.cell(row=2, column=6, value='{"value": 7850}')

        ws_props.cell(row=3, column=1, value='AISI 4340')
        ws_props.cell(row=3, column=2, value='thermal_conductivity')
        ws_props.cell(row=3, column=3, value='curve')
        ws_props.cell(row=3, column=4, value='W/(m·K)')
        ws_props.cell(row=3, column=5, value='temperature')
        ws_props.cell(row=3, column=6, value='{"temperature": [20, 200, 400], "value": [44, 42, 38]}')

        # Sheet 3: Phase Diagrams
        ws_diagrams = wb.create_sheet('Phase Diagrams')
        headers = ['designation', 'diagram_type', 'Ac1', 'Ac3', 'Ms', 'Mf']
        for col, header in enumerate(headers, 1):
            cell = ws_diagrams.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill

        # Example row
        ws_diagrams.cell(row=2, column=1, value='AISI 4340')
        ws_diagrams.cell(row=2, column=2, value='CCT')
        ws_diagrams.cell(row=2, column=3, value=727)
        ws_diagrams.cell(row=2, column=4, value=780)
        ws_diagrams.cell(row=2, column=5, value=320)
        ws_diagrams.cell(row=2, column=6, value=180)

        # Auto-adjust column widths
        for ws in [ws_grades, ws_props, ws_diagrams]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
