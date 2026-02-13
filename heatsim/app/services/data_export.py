"""Data export utilities for simulation results and measured data.

Exports to CSV and Excel (openpyxl) formats.
"""
import csv
import io
from typing import List

import numpy as np

from app.models.simulation import Simulation, SimulationResult


class DataExporter:
    """Export simulation and measured data to CSV / Excel."""

    @staticmethod
    def export_simulation_csv(sim: Simulation) -> str:
        """Export full-cycle temperature results as CSV.

        Columns: time_s, center_C, one_third_C, two_thirds_C, surface_C
        """
        cycle = sim.results.filter_by(result_type='full_cycle').first()
        if not cycle:
            return ''

        times = cycle.time_array
        data = cycle.data_dict  # {center: [...], one_third: [...], ...}

        positions = ['center', 'one_third', 'two_thirds', 'surface']
        header = ['time_s'] + [f'{p}_C' for p in positions]

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(header)

        # value_array holds center temps; data_dict may have multi-position
        center = data.get('center', cycle.value_array)
        one_third = data.get('one_third', [])
        two_thirds = data.get('two_thirds', [])
        surface = data.get('surface', [])

        for i, t in enumerate(times):
            row = [f'{t:.3f}']
            for arr in [center, one_third, two_thirds, surface]:
                if i < len(arr):
                    row.append(f'{arr[i]:.2f}')
                else:
                    row.append('')
            writer.writerow(row)

        return buf.getvalue()

    @staticmethod
    def export_simulation_excel(sim: Simulation) -> bytes:
        """Export simulation data as multi-sheet Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        bold = Font(bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # --- Config sheet ---
        ws = wb.active
        ws.title = 'Configuration'
        config_rows = [
            ('Simulation', sim.name),
            ('Steel Grade', sim.steel_grade.designation),
            ('Geometry', sim.geometry_label),
            ('Status', sim.status),
            ('Created', str(sim.created_at)),
        ]
        ht = sim.ht_config
        if ht.get('heating', {}).get('enabled'):
            h = ht['heating']
            config_rows.append(('Heating Temp (°C)', h.get('target_temperature', '')))
            config_rows.append(('Heating Hold (min)', h.get('hold_time', '')))
        q = ht.get('quenching', {})
        if q:
            config_rows.append(('Quench Media', q.get('media', '')))
            config_rows.append(('Quench Temp (°C)', q.get('media_temperature', '')))
            config_rows.append(('Agitation', q.get('agitation', '')))
        if ht.get('tempering', {}).get('enabled'):
            t = ht['tempering']
            config_rows.append(('Tempering Temp (°C)', t.get('temperature', '')))
            config_rows.append(('Tempering Hold (min)', t.get('hold_time', '')))
        for r, (k, v) in enumerate(config_rows, 1):
            ws.cell(r, 1, k).font = bold
            ws.cell(r, 2, v)
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 30

        # --- Temperature sheet ---
        cycle = sim.results.filter_by(result_type='full_cycle').first()
        if cycle:
            ws_temp = wb.create_sheet('Temperature')
            positions = ['center', 'one_third', 'two_thirds', 'surface']
            headers = ['time_s'] + [f'{p}_C' for p in positions]
            for c, h in enumerate(headers, 1):
                cell = ws_temp.cell(1, c, h)
                cell.font = bold
                cell.fill = header_fill

            times = cycle.time_array
            data = cycle.data_dict
            center = data.get('center', cycle.value_array)
            cols = [center, data.get('one_third', []), data.get('two_thirds', []), data.get('surface', [])]
            for i, t in enumerate(times):
                ws_temp.cell(i + 2, 1, round(t, 3))
                for j, arr in enumerate(cols):
                    if i < len(arr):
                        ws_temp.cell(i + 2, j + 2, round(arr[i], 2))

        # --- Phases sheet ---
        phase_result = sim.results.filter_by(result_type='phase_fraction').first()
        if phase_result:
            ws_ph = wb.create_sheet('Phases')
            phases_data = phase_result.phases_dict
            # phases_data: {martensite: val, bainite: val, pearlite: val, ferrite: val, ...}
            # Or could be time-series: {time: [...], martensite: [...], ...}
            if 'time' in phases_data:
                phase_names = [k for k in phases_data if k != 'time']
                headers = ['time_s'] + phase_names
                for c, h in enumerate(headers, 1):
                    cell = ws_ph.cell(1, c, h)
                    cell.font = bold
                    cell.fill = header_fill
                times_p = phases_data['time']
                for i, t in enumerate(times_p):
                    ws_ph.cell(i + 2, 1, round(t, 3))
                    for j, pn in enumerate(phase_names):
                        vals = phases_data.get(pn, [])
                        if i < len(vals):
                            ws_ph.cell(i + 2, j + 2, round(vals[i], 4))
            else:
                # Single-point summary
                ws_ph.cell(1, 1, 'Phase').font = bold
                ws_ph.cell(1, 2, 'Fraction').font = bold
                for r, (pn, pv) in enumerate(phases_data.items(), 2):
                    ws_ph.cell(r, 1, pn)
                    ws_ph.cell(r, 2, round(float(pv), 4) if isinstance(pv, (int, float)) else pv)

        # --- Hardness sheet ---
        hardness_result = sim.results.filter_by(result_type='hardness_prediction').first()
        if hardness_result:
            ws_hd = wb.create_sheet('Hardness')
            hd = hardness_result.data_dict
            ws_hd.cell(1, 1, 'Property').font = bold
            ws_hd.cell(1, 2, 'Value').font = bold
            row = 2
            for k, v in hd.items():
                ws_hd.cell(row, 1, k)
                if isinstance(v, (int, float)):
                    ws_hd.cell(row, 2, round(v, 2))
                else:
                    ws_hd.cell(row, 2, str(v))
                row += 1

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    @staticmethod
    def export_measured_csv(md) -> str:
        """Export measured data as CSV with interpolated common time grid.

        Parameters
        ----------
        md : MeasuredData
            Measured data object
        """
        channels = md.channels  # {TC1: [temps], TC2: [temps], ...}
        channel_times = md.channel_times  # {TC1: [times], ...}
        if not channels:
            return ''

        # Build common time grid from all channel times
        all_times = set()
        for ch, times in channel_times.items():
            all_times.update(times)
        common_times = sorted(all_times)
        if not common_times:
            return ''

        common_times_arr = np.array(common_times)
        channel_names = sorted(channels.keys())

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(['time_s'] + channel_names)

        # Interpolate each channel to common grid
        interp_data = {}
        for ch in channel_names:
            ch_t = np.array(channel_times.get(ch, md.times))
            ch_v = np.array(channels[ch])
            if len(ch_t) > 1 and len(ch_v) > 1:
                interp_data[ch] = np.interp(common_times_arr, ch_t, ch_v)
            else:
                interp_data[ch] = np.full(len(common_times_arr), np.nan)

        for i, t in enumerate(common_times):
            row = [f'{t:.3f}']
            for ch in channel_names:
                v = interp_data[ch][i]
                row.append(f'{v:.2f}' if not np.isnan(v) else '')
            writer.writerow(row)

        return buf.getvalue()

    @staticmethod
    def export_comparison_csv(sim: Simulation) -> str:
        """Export comparison metrics between simulation and measured data.

        One row per measured dataset linked to the simulation.
        """
        from app.services.comparison_service import ComparisonService

        measured_list = sim.measured_data.all()
        if not measured_list:
            return ''

        cycle = sim.results.filter_by(result_type='full_cycle').first()
        if not cycle:
            return ''

        sim_times = np.array(cycle.time_array)
        sim_temps = np.array(cycle.data_dict.get('center', cycle.value_array))

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            'measured_name', 'process_step', 'rms_error_C', 'peak_temp_diff_C',
            'r_squared', 'max_abs_error_C', 'time_offset_s', 'rating'
        ])

        for md in measured_list:
            # Use first channel for comparison
            ch_names = md.available_channels
            if not ch_names:
                continue
            ch = ch_names[0]
            meas_times = np.array(md.get_channel_times(ch))
            meas_temps = np.array(md.get_channel_data(ch))
            if len(meas_times) < 2 or len(meas_temps) < 2:
                continue

            metrics = ComparisonService.compare(sim_times, sim_temps, meas_times, meas_temps)
            writer.writerow([
                md.name, md.process_step or 'full',
                f'{metrics.rms_error:.2f}', f'{metrics.peak_temp_diff:.2f}',
                f'{metrics.r_squared:.4f}', f'{metrics.max_abs_error:.2f}',
                f'{metrics.time_offset:.2f}', metrics.rating
            ])

        return buf.getvalue()

    @staticmethod
    def export_sweep_summary_csv(sim_ids: List[int]) -> str:
        """Export summary table for a parameter sweep (multiple simulations).

        Columns: sim_name, steel_grade, geometry, quench_media, t8_5_s,
                 hardness_HV, martensite_pct.
        """
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            'simulation', 'steel_grade', 'geometry', 'quench_media',
            't8_5_s', 'hardness_HV', 'martensite_pct'
        ])

        for sid in sim_ids:
            sim = Simulation.query.get(sid)
            if not sim:
                continue

            cycle = sim.results.filter_by(result_type='full_cycle').first()
            t85 = cycle.t_800_500 if cycle else None

            hardness_result = sim.results.filter_by(result_type='hardness_prediction').first()
            hv = hardness_result.data_dict.get('surface_HV') if hardness_result else None
            if hv is None and hardness_result:
                hv = hardness_result.data_dict.get('center_HV')

            phase_result = sim.results.filter_by(result_type='phase_fraction').first()
            martensite = None
            if phase_result:
                pd = phase_result.phases_dict
                if 'martensite' in pd:
                    val = pd['martensite']
                    martensite = val if isinstance(val, (int, float)) else None

            q_media = sim.ht_config.get('quenching', {}).get('media', '')

            writer.writerow([
                sim.name, sim.steel_grade.designation, sim.geometry_label,
                q_media,
                f'{t85:.2f}' if t85 else '',
                f'{hv:.1f}' if hv else '',
                f'{martensite:.4f}' if martensite is not None else ''
            ])

        return buf.getvalue()
