"""Excel export functionality for reports.

Exports comparison reports, project reports, and financial statements to Excel.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

from app.models import FactProjectMonthly
from .report_analysis import (
    generate_kpi_report1,
    generate_kpi_report2,
    generate_summary_sheet
)


def _style_header(ws, row_num, col_count):
    """Apply header styling to a row."""
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')


def _style_currency(ws, col_letter, start_row, end_row):
    """Apply currency formatting to a column."""
    for row in range(start_row, end_row + 1):
        cell = ws[f'{col_letter}{row}']
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='right')


def _write_comparison_sheet(ws, df, cols, headers):
    """Write a comparison sheet with selected columns and custom headers."""
    export_cols = [c for c in cols if c in df.columns]
    for c_idx, header in enumerate(headers[:len(export_cols)], 1):
        ws.cell(row=1, column=c_idx, value=header)
    _style_header(ws, 1, len(export_cols))

    for r_idx, (_, row) in enumerate(df[export_cols].iterrows(), 2):
        for c_idx, col in enumerate(export_cols, 1):
            ws.cell(row=r_idx, column=c_idx, value=row[col])

    # Totals row
    total_row = len(df) + 2
    ws.cell(row=total_row, column=1, value='TOTAL')
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    for c_idx, col in enumerate(export_cols, 1):
        if c_idx > 2:  # skip project_number and project_name
            total = df[col].sum()
            ws.cell(row=total_row, column=c_idx, value=total)
            ws.cell(row=total_row, column=c_idx).font = Font(bold=True)
            ws.cell(row=total_row, column=c_idx).number_format = '#,##0'


def export_comparison_to_excel(comparison_result: dict) -> BytesIO:
    """Export comparison results to Excel workbook.

    Args:
        comparison_result: Dictionary from compare_two_months()

    Returns:
        BytesIO buffer containing Excel file
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    current_date = comparison_result.get('current_date', 'Current')
    previous_date = comparison_result.get('previous_date', 'Previous')

    # Sheet 1: Summary
    ws_summary = wb.create_sheet('Summary')
    summary_curr = comparison_result['summary_curr']
    summary_prev = comparison_result['summary_prev']

    # Write summary headers
    ws_summary['A1'] = 'Metric'
    ws_summary['B1'] = f'Current ({current_date})'
    ws_summary['C1'] = f'Previous ({previous_date})'
    ws_summary['D1'] = 'Change'
    _style_header(ws_summary, 1, 4)

    # Summary data
    metrics = [
        ('Project Count', 'project_count'),
        ('Total Revenue', 'total_revenue'),
        ('Total Cost', 'total_cost'),
        ('Gross Margin', 'total_gross_margin'),
        ('Total Accrued', 'total_accrued'),
        ('Total Contingency', 'total_contingency'),
    ]

    for i, (label, key) in enumerate(metrics, start=2):
        curr_val = summary_curr.get(key, 0)
        prev_val = summary_prev.get(key, 0)
        ws_summary[f'A{i}'] = label
        ws_summary[f'B{i}'] = curr_val
        ws_summary[f'C{i}'] = prev_val
        ws_summary[f'D{i}'] = curr_val - prev_val

    detail_df = comparison_result['detail']

    # Sheet 2: Revenue Comparison
    ws_rev = wb.create_sheet('Revenue Comparison')
    if not detail_df.empty:
        rev_cols = [
            'project_number', 'project_name',
            'total_income_curr', 'total_income_prev', 'delta_total_income',
            'accrued_income_curr', 'accrued_income_prev', 'delta_accrued_income',
            'remaining_income_curr', 'remaining_income_prev', 'delta_remaining_income',
        ]
        rev_headers = [
            'Project', 'Name',
            f'Total Income ({current_date})', f'Total Income ({previous_date})', 'Delta',
            f'Accrued ({current_date})', f'Accrued ({previous_date})', 'Delta',
            f'Remaining ({current_date})', f'Remaining ({previous_date})', 'Delta',
        ]
        _write_comparison_sheet(ws_rev, detail_df, rev_cols, rev_headers)

    # Sheet 3: Profit Comparison
    ws_profit = wb.create_sheet('Profit Comparison')
    if not detail_df.empty:
        profit_cols = [
            'project_number', 'project_name',
            'total_income_curr', 'total_income_prev', 'delta_total_income',
            'total_cost_curr', 'total_cost_prev', 'delta_total_cost',
            'project_profit_curr', 'project_profit_prev', 'delta_project_profit',
        ]
        profit_headers = [
            'Project', 'Name',
            f'Income ({current_date})', f'Income ({previous_date})', 'Delta',
            f'Cost ({current_date})', f'Cost ({previous_date})', 'Delta',
            f'Profit ({current_date})', f'Profit ({previous_date})', 'Delta',
        ]
        _write_comparison_sheet(ws_profit, detail_df, profit_cols, profit_headers)

    # Sheet 4: Order Book Comparison
    ws_ob = wb.create_sheet('Order Book Comparison')
    if not detail_df.empty:
        ob_cols = [
            'project_number', 'project_name',
            'remaining_income_curr', 'remaining_income_prev', 'delta_remaining_income',
            'remaining_cost_curr', 'remaining_cost_prev', 'delta_remaining_cost',
            'completion_curr', 'completion_prev', 'delta_completion',
        ]
        ob_headers = [
            'Project', 'Name',
            f'Rem. Income ({current_date})', f'Rem. Income ({previous_date})', 'Delta',
            f'Rem. Cost ({current_date})', f'Rem. Cost ({previous_date})', 'Delta',
            f'Completion ({current_date})', f'Completion ({previous_date})', 'Delta',
        ]
        _write_comparison_sheet(ws_ob, detail_df, ob_cols, ob_headers)

    # Sheet 5: Booking Comparison
    ws_book = wb.create_sheet('Booking Comparison')
    if not detail_df.empty:
        book_cols = [
            'project_number', 'project_name',
            'accrued_income_curr', 'accrued_income_prev', 'delta_accrued_income',
            'contingency_curr', 'contingency_prev', 'delta_contingency',
            'net_booking_curr', 'net_booking_prev', 'delta_net_booking',
        ]
        book_headers = [
            'Project', 'Name',
            f'Accrued ({current_date})', f'Accrued ({previous_date})', 'Delta',
            f'Contingency ({current_date})', f'Contingency ({previous_date})', 'Delta',
            f'Net Booking ({current_date})', f'Net Booking ({previous_date})', 'Delta',
        ]
        _write_comparison_sheet(ws_book, detail_df, book_cols, book_headers)

    # Adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_project_report(closing_date: str) -> BytesIO:
    """Export project report for a closing date.

    Args:
        closing_date: Closing date (YYYY-MM-DD)

    Returns:
        BytesIO buffer containing Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Project Report'

    # Get projects
    projects = FactProjectMonthly.get_by_closing_date(closing_date)

    if not projects:
        ws['A1'] = 'No data available'
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # Headers
    headers = [
        'Project', 'Name', 'Customer',
        'Expected Income', 'Expected Cost',
        'Actual Income', 'Actual Cost',
        'Remaining Income', 'Remaining Cost',
        'Total Income', 'Total Cost',
        'Completion %', 'Profit Margin %',
        'Accrued Income', 'Contingency',
        'Project Profit'
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, 1, len(headers))

    # Data
    for row_num, p in enumerate(projects, start=2):
        ws.cell(row=row_num, column=1, value=p.project_number)
        ws.cell(row=row_num, column=2, value=p.project_name)
        ws.cell(row=row_num, column=3, value=p.customer_name)
        ws.cell(row=row_num, column=4, value=p.expected_income or 0)
        ws.cell(row=row_num, column=5, value=p.expected_cost or 0)
        ws.cell(row=row_num, column=6, value=p.actual_income_cur or 0)
        ws.cell(row=row_num, column=7, value=p.actual_cost_cur or 0)
        ws.cell(row=row_num, column=8, value=p.remaining_income or 0)
        ws.cell(row=row_num, column=9, value=p.remaining_cost or 0)
        ws.cell(row=row_num, column=10, value=p.total_income_cur or 0)
        ws.cell(row=row_num, column=11, value=p.total_cost_cur or 0)
        ws.cell(row=row_num, column=12, value=(p.completion_cur1 or 0) * 100)
        ws.cell(row=row_num, column=13, value=(p.profit_margin_cur or 0) * 100)
        ws.cell(row=row_num, column=14, value=p.accrued_income_cur or 0)
        ws.cell(row=row_num, column=15, value=p.contingency_cur or 0)
        ws.cell(row=row_num, column=16, value=p.project_profit or 0)

    # Totals row
    total_row = len(projects) + 2
    ws.cell(row=total_row, column=1, value='TOTAL')
    ws.cell(row=total_row, column=1).font = Font(bold=True)

    for col in [4, 5, 6, 7, 8, 9, 10, 11, 14, 15, 16]:
        ws.cell(row=total_row, column=col,
                value=f'=SUM({chr(64+col)}2:{chr(64+col)}{total_row-1})')
        ws.cell(row=total_row, column=col).font = Font(bold=True)

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 20)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_financial_statements(pnl_data: dict, bs_data: dict) -> BytesIO:
    """Export P&L and Balance Sheet to Excel.

    Args:
        pnl_data: P&L data dictionary
        bs_data: Balance Sheet data dictionary

    Returns:
        BytesIO buffer containing Excel file
    """
    wb = Workbook()
    wb.remove(wb.active)

    # P&L Sheet
    ws_pnl = wb.create_sheet('P&L Statement')
    ws_pnl['A1'] = f"Profit & Loss Statement"
    ws_pnl['A2'] = f"{pnl_data.get('start_date', '')} to {pnl_data.get('end_date', '')}"
    ws_pnl['A1'].font = Font(bold=True, size=14)

    row = 4
    for line in pnl_data.get('lines', []):
        indent = '  ' * line.get('level', 0)
        ws_pnl[f'A{row}'] = indent + line['line']
        if line.get('amount') is not None:
            ws_pnl[f'B{row}'] = line['amount']
            ws_pnl[f'B{row}'].number_format = '#,##0'
        if line.get('bold'):
            ws_pnl[f'A{row}'].font = Font(bold=True)
            ws_pnl[f'B{row}'].font = Font(bold=True)
        row += 1

    # Balance Sheet
    ws_bs = wb.create_sheet('Balance Sheet')
    ws_bs['A1'] = f"Balance Sheet"
    ws_bs['A2'] = f"As at {bs_data.get('closing_date', '')}"
    ws_bs['A1'].font = Font(bold=True, size=14)

    row = 4
    # Assets
    for line in bs_data.get('asset_lines', []):
        indent = '  ' * line.get('level', 0)
        ws_bs[f'A{row}'] = indent + line['line']
        if line.get('amount') is not None:
            ws_bs[f'B{row}'] = line['amount']
            ws_bs[f'B{row}'].number_format = '#,##0'
        if line.get('bold'):
            ws_bs[f'A{row}'].font = Font(bold=True)
            ws_bs[f'B{row}'].font = Font(bold=True)
        row += 1

    row += 1
    # Liabilities & Equity
    for line in bs_data.get('liability_lines', []):
        indent = '  ' * line.get('level', 0)
        ws_bs[f'A{row}'] = indent + line['line']
        if line.get('amount') is not None:
            ws_bs[f'B{row}'] = line['amount']
            ws_bs[f'B{row}'].number_format = '#,##0'
        if line.get('bold'):
            ws_bs[f'A{row}'].font = Font(bold=True)
            ws_bs[f'B{row}'].font = Font(bold=True)
        row += 1

    # Adjust column widths
    for ws in wb.worksheets:
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_management_reports(generator) -> BytesIO:
    """Export management reports (Report1, Report2, Report3) to Excel.

    Args:
        generator: ProjectReportGenerator instance

    Returns:
        BytesIO buffer containing Excel file
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Summary sheet
    ws_summary = wb.create_sheet('Summary')
    summary = generator.generate_summary()

    ws_summary['A1'] = 'Project Accrued Income Report'
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A2'] = f"Closing Date: {summary['closing_date']}"

    row = 4
    summary_items = [
        ('Project Count', summary['project_count'], ''),
        ('Accrued Revenue', summary['accrued_revenue'], '#,##0'),
        ('Orderbook', summary['orderbook'], '#,##0'),
        ('Total Revenue', summary['total_revenue'], '#,##0'),
        ('Total COGS', summary['total_cogs'], '#,##0'),
        ('Gross Profit', summary['gross_profit'], '#,##0'),
        ('Gross Margin %', summary['gross_margin_pct'] * 100, '0.0%'),
        ('Total Contingency', summary['contingency'], '#,##0'),
    ]

    for label, value, fmt in summary_items:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        if fmt:
            ws_summary[f'B{row}'].number_format = fmt
        row += 1

    # Report1 sheet
    ws_r1 = wb.create_sheet('Report1')
    report1 = generator.generate_report1()

    headers_r1 = [
        'Project', 'Name',
        'Curr Revenue', 'Curr COGS', 'Curr GM', 'Curr GM%',
        'YTD Revenue', 'YTD COGS', 'YTD GM', 'YTD GM%',
        'PTD Revenue', 'PTD COGS', 'PTD GM', 'PTD GM%'
    ]

    for col, header in enumerate(headers_r1, 1):
        ws_r1.cell(row=1, column=col, value=header)
    _style_header(ws_r1, 1, len(headers_r1))

    for row_num, r in enumerate(report1, start=2):
        ws_r1.cell(row=row_num, column=1, value=r.project_number)
        ws_r1.cell(row=row_num, column=2, value=r.project_name)
        ws_r1.cell(row=row_num, column=3, value=r.current_revenue)
        ws_r1.cell(row=row_num, column=4, value=r.current_cogs)
        ws_r1.cell(row=row_num, column=5, value=r.current_gm)
        ws_r1.cell(row=row_num, column=6, value=r.current_gm_pct)
        ws_r1.cell(row=row_num, column=7, value=r.ytd_revenue)
        ws_r1.cell(row=row_num, column=8, value=r.ytd_cogs)
        ws_r1.cell(row=row_num, column=9, value=r.ytd_gm)
        ws_r1.cell(row=row_num, column=10, value=r.ytd_gm_pct)
        ws_r1.cell(row=row_num, column=11, value=r.ptd_revenue)
        ws_r1.cell(row=row_num, column=12, value=r.ptd_cogs)
        ws_r1.cell(row=row_num, column=13, value=r.ptd_gm)
        ws_r1.cell(row=row_num, column=14, value=r.ptd_gm_pct)

        # Format percentages
        ws_r1.cell(row=row_num, column=6).number_format = '0.0%'
        ws_r1.cell(row=row_num, column=10).number_format = '0.0%'
        ws_r1.cell(row=row_num, column=14).number_format = '0.0%'

    # Report2 sheet
    ws_r2 = wb.create_sheet('Report2')
    report2 = generator.generate_report2()

    headers_r2 = [
        'Project', 'Name',
        'TCV/Revenue', 'COGS', 'GM', 'GM%', 'PoC', 'Contingency',
        'Delta Revenue', 'Delta COGS', 'Delta GM', 'Currency Effect'
    ]

    for col, header in enumerate(headers_r2, 1):
        ws_r2.cell(row=1, column=col, value=header)
    _style_header(ws_r2, 1, len(headers_r2))

    for row_num, r in enumerate(report2, start=2):
        ws_r2.cell(row=row_num, column=1, value=r.project_number)
        ws_r2.cell(row=row_num, column=2, value=r.project_name)
        ws_r2.cell(row=row_num, column=3, value=r.tcv_revenue)
        ws_r2.cell(row=row_num, column=4, value=r.cogs)
        ws_r2.cell(row=row_num, column=5, value=r.gm)
        ws_r2.cell(row=row_num, column=6, value=r.gm_pct)
        ws_r2.cell(row=row_num, column=7, value=r.poc)
        ws_r2.cell(row=row_num, column=8, value=r.contingency)
        ws_r2.cell(row=row_num, column=9, value=r.delta_revenue)
        ws_r2.cell(row=row_num, column=10, value=r.delta_cogs)
        ws_r2.cell(row=row_num, column=11, value=r.delta_gm)
        ws_r2.cell(row=row_num, column=12, value=r.currency_effect)

        # Format percentages
        ws_r2.cell(row=row_num, column=6).number_format = '0.0%'
        ws_r2.cell(row=row_num, column=7).number_format = '0.0%'

    # Report3 sheet
    ws_r3 = wb.create_sheet('Report3')
    report3 = generator.generate_report3()

    headers_r3 = [
        'Project', 'Name',
        'OB Order Backlog', 'Order Intake', 'Revenue', 'EB Order Backlog',
        'GM%', 'Production Start', 'Delivery'
    ]

    for col, header in enumerate(headers_r3, 1):
        ws_r3.cell(row=1, column=col, value=header)
    _style_header(ws_r3, 1, len(headers_r3))

    for row_num, r in enumerate(report3, start=2):
        ws_r3.cell(row=row_num, column=1, value=r.project_number)
        ws_r3.cell(row=row_num, column=2, value=r.project_name)
        ws_r3.cell(row=row_num, column=3, value=r.ob_order_backlog)
        ws_r3.cell(row=row_num, column=4, value=r.order_intake)
        ws_r3.cell(row=row_num, column=5, value=r.revenue)
        ws_r3.cell(row=row_num, column=6, value=r.eb_order_backlog)
        ws_r3.cell(row=row_num, column=7, value=r.gm_pct)
        ws_r3.cell(row=row_num, column=8, value=r.production_start or '')
        ws_r3.cell(row=row_num, column=9, value=r.delivery or '')

        # Format percentage
        ws_r3.cell(row=row_num, column=7).number_format = '0.0%'

    # Adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 20)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
