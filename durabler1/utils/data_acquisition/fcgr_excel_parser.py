"""
FCGR Excel file parser for MTS TestSuite Analysis Reports.

Imports specimen dimensions, material properties, test parameters,
and FCG results from Excel files exported by MTS TestSuite.
"""

from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any
import numpy as np


@dataclass
class FCGRUserInputs:
    """
    Container for FCGR user input data imported from Excel.

    Parameters
    ----------
    specimen_id : str
        Specimen identifier (Test Run name)
    specimen_type : str
        'C(T)' or 'M(T)'
    W : float
        Specimen width (mm)
    B : float
        Specimen thickness (mm)
    B_n : float
        Net thickness at side grooves (mm)
    a_0 : float
        Initial notch length (mm)
    notch_height : float
        Notch height h (mm)
    yield_strength : float
        0.2% yield strength (MPa)
    ultimate_strength : float
        Ultimate tensile strength (MPa)
    youngs_modulus : float
        Young's modulus (GPa)
    poissons_ratio : float
        Poisson's ratio
    test_temperature : float
        Test temperature (C)
    control_mode : str
        'Load Control' or 'Delta-K Control'
    load_ratio : float
        R ratio (Pmin/Pmax)
    frequency : float
        Test frequency (Hz)
    compliance_coefficients : List[float]
        Compliance calibration coefficients C0-C5
    k_calibration_coefficients : List[float]
        K calibration coefficients
    precrack_measurements : List[float]
        Precrack crack size measurements (mm)
    final_crack_measurements : List[float]
        Final crack size measurements (mm)
    paris_law_C : float
        Paris law coefficient C
    paris_law_m : float
        Paris law exponent m
    fcgr_results : Dict[str, any]
        FCG results from MTS analysis
    """
    specimen_id: str
    specimen_type: str
    W: float
    B: float
    B_n: float
    a_0: float
    notch_height: float = 0.0
    yield_strength: float = 0.0
    ultimate_strength: float = 0.0
    youngs_modulus: float = 210.0
    poissons_ratio: float = 0.3
    test_temperature: float = 23.0
    control_mode: str = "Load Control"
    load_ratio: float = 0.1
    frequency: float = 10.0
    wave_shape: str = "Sine"
    normalized_k_gradient: float = 0.0
    compliance_coefficients: List[float] = field(default_factory=list)
    k_calibration_coefficients: List[float] = field(default_factory=list)
    precrack_measurements: List[float] = field(default_factory=list)
    precrack_final_size: float = 0.0
    final_crack_measurements: List[float] = field(default_factory=list)
    final_crack_size: float = 0.0
    paris_law_C: float = 0.0
    paris_law_m: float = 0.0
    threshold_delta_K: float = 0.0
    total_cycles: int = 0
    fcgr_results: Dict[str, Any] = field(default_factory=dict)
    material: str = ""

    @property
    def precrack_average(self) -> float:
        """Calculate average precrack length."""
        if not self.precrack_measurements:
            return self.a_0
        if len(self.precrack_measurements) == 5:
            a = self.precrack_measurements
            return (0.5 * a[0] + a[1] + a[2] + a[3] + 0.5 * a[4]) / 4
        return np.mean(self.precrack_measurements)

    @property
    def final_crack_average(self) -> float:
        """Calculate average final crack length."""
        if not self.final_crack_measurements:
            return self.final_crack_size
        if len(self.final_crack_measurements) == 5:
            a = self.final_crack_measurements
            return (0.5 * a[0] + a[1] + a[2] + a[3] + 0.5 * a[4]) / 4
        return np.mean(self.final_crack_measurements)

    @property
    def a_W_ratio(self) -> float:
        """Calculate initial a/W ratio."""
        if self.W == 0:
            return 0.0
        a = self.precrack_average if self.precrack_measurements else self.a_0
        return a / self.W


def parse_fcgr_excel(filepath: Path) -> FCGRUserInputs:
    """
    Parse FCGR user input data from MTS Excel Analysis Report.

    The Excel file contains specimen dimensions, material properties,
    test parameters, and FCG analysis results.

    Parameters
    ----------
    filepath : Path
        Path to Excel file (.xlsx)

    Returns
    -------
    FCGRUserInputs
        Parsed user input data

    Raises
    ------
    FileNotFoundError
        If file does not exist
    ValueError
        If required data cannot be found
    """
    from openpyxl import load_workbook

    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    try:
        wb = load_workbook(filepath, data_only=True)
        if 'Data' in wb.sheetnames:
            ws = wb['Data']
        else:
            ws = wb.active
    except Exception as e:
        raise ValueError(f"Failed to read Excel file: {e}")

    # Build a dictionary of all label-value pairs
    data_dict = {}
    for row in ws.iter_rows(min_row=1, max_row=200, values_only=True):
        if row and len(row) >= 9:
            label = row[2] if len(row) > 2 else None  # Column C
            value = row[8] if len(row) > 8 else None  # Column I
            if label and value is not None:
                data_dict[str(label).strip()] = value

    def parse_value(val_str, default=0.0):
        """Extract numeric value from string like '83.1000 mm'."""
        if val_str is None:
            return default
        val_str = str(val_str)
        parts = val_str.split()
        if parts:
            try:
                # Handle scientific notation
                val = parts[0].replace(',', '.')
                return float(val)
            except ValueError:
                return default
        return default

    def get_value(keys, default=0.0):
        """Get value from data_dict using multiple possible keys."""
        for key in keys:
            if key in data_dict:
                return parse_value(data_dict[key], default)
        return default

    def get_string(keys, default=''):
        """Get string value from data_dict."""
        for key in keys:
            if key in data_dict:
                val = data_dict[key]
                if val:
                    return str(val)
        return default

    # Extract specimen info
    specimen_id = get_string(['Name of the Test Run', 'Specimen Name'], 'Unknown')
    geometry_type = get_string(['Geometry Type'], 'C(T)')

    # Determine specimen type
    geom_lower = geometry_type.lower()
    if 'c(t)' in geom_lower or 'ct' in geom_lower or 'compact' in geom_lower or 'llc' in geom_lower:
        specimen_type = 'C(T)'
    elif 'm(t)' in geom_lower or 'mt' in geom_lower or 'middle' in geom_lower:
        specimen_type = 'M(T)'
    else:
        specimen_type = 'C(T)'

    # Specimen dimensions
    W = get_value(['Width (W)'], 50.0)
    B = get_value(['Thickness (B)'], 12.5)
    B_n = get_value(['Net Thickness (Bn)'], B)
    a_0 = get_value(['Notch Length (a0)'], 10.0)
    notch_height = get_value(['Notch Height (h)'], 0.0)

    # Material properties (MTS uses kN/mm² for stress, convert to MPa)
    E_raw = get_value(['Elastic Modulus (E)', 'User Defined Elastic Modulus'], 210.0)
    youngs_modulus = E_raw  # kN/mm² = GPa

    yield_raw = get_value(['Yield Strength'], 0.5)
    yield_strength = yield_raw * 1000  # kN/mm² to MPa

    uts_raw = get_value(['Ultimate Tensile Strength'], 0.6)
    ultimate_strength = uts_raw * 1000  # kN/mm² to MPa

    poissons_ratio = get_value(["Poisson's Ratio"], 0.3)
    test_temperature = get_value(['Valid at Temperature'], 23.0)

    # Test parameters
    control_mode = get_string(['FCG Control Mode'], 'Constant Load')
    if 'constant' in control_mode.lower() or 'load' in control_mode.lower():
        control_mode = 'Load Control'
    else:
        control_mode = 'Delta-K Control'

    load_ratio = get_value(['FCG Load Ratio'], 0.1)
    frequency = get_value(['FCG Frequency'], 10.0)
    wave_shape = get_string(['FCG Wave Shape'], 'Sine')
    normalized_k_gradient = get_value(['Normalized K Gradient (C)'], 0.0)

    # Compliance coefficients (C0-C5)
    compliance_coefficients = []
    for i in range(6):
        coef = get_value([f'Compliance Coefficient C{i}'], 0.0)
        compliance_coefficients.append(coef)

    # K calibration coefficients
    k_calibration_coefficients = []
    for i in range(5):
        coef = get_value([f'K Calibration Coefficient {i}'], 0.0)
        k_calibration_coefficients.append(coef)

    # Precrack measurements
    precrack_measurements = []
    for i in range(1, 6):
        val = get_value([f'Precrack Crack {i}'], 0.0)
        if val > 0:
            precrack_measurements.append(val)

    precrack_final_size = get_value(['Precrack Final Crack Size'], a_0)

    # Final crack measurements
    final_crack_measurements = []
    for i in range(1, 6):
        val = get_value([f'FCG Crack {i}'], 0.0)
        if val > 0:
            final_crack_measurements.append(val)

    final_crack_size = get_value(['FCG Final Crack Size'], 0.0)
    total_cycles = int(get_value(['FCG Cycles Completed'], 0))

    # Paris law results
    paris_law_C = get_value(['C Coefficient'], 0.0)
    paris_law_m = get_value(['m Threshold Coefficient', 'm Coefficient'], 0.0)
    threshold_delta_K = get_value(['Apparent Delta-K Threshold'], 0.0)

    # FCG results summary
    fcgr_results = {
        'initial_K_max': get_value(['FCG: initial Kmax'], 0.0),
        'final_K_max': get_value(['FCG Final K Maximum'], 0.0),
        'final_K': get_value(['FCG Final K'], 0.0),
        'final_P_max': get_value(['FCG Final P Maximum'], 0.0),
        'final_P': get_value(['FCG Final P'], 0.0),
        'reason_for_end': get_string(['FCG Reason for End'], ''),
        'crack_limit': get_value(['FCG Final Crack Limit'], 0.0),
        'growth_rate_limit': get_value(['FCG Crack Growth Rate Limit'], 0.0),
        'da_dN_curve_fit': get_string(['FCG da/dN Curve Fit'], 'Secant'),
        'polynomial_fit_number': int(get_value(['Polynomial Fit Number'], 3)),
        # Validity checks
        'notch_valid': get_string(['Is Notch Length >= 0.2W?'], ''),
        'precrack_extension_valid': get_string(['Is precrack extension from notch >= 0.10B, h or 1 mm?'], ''),
        'precrack_kmax_valid': get_string(['Is final precrack Kmax <= initial test Kmax?'], ''),
        'precrack_diff_valid': get_string(['Is precrack difference <= 0.025W or 0.25B'], ''),
        'final_diff_valid': get_string(['Is final crack difference <= 0.025W or 0.25B'], ''),
    }

    return FCGRUserInputs(
        specimen_id=str(specimen_id),
        specimen_type=specimen_type,
        W=W,
        B=B,
        B_n=B_n if B_n else B,
        a_0=a_0,
        notch_height=notch_height,
        yield_strength=yield_strength,
        ultimate_strength=ultimate_strength,
        youngs_modulus=youngs_modulus,
        poissons_ratio=poissons_ratio,
        test_temperature=test_temperature,
        control_mode=control_mode,
        load_ratio=load_ratio,
        frequency=frequency,
        wave_shape=wave_shape,
        normalized_k_gradient=normalized_k_gradient,
        compliance_coefficients=compliance_coefficients,
        k_calibration_coefficients=k_calibration_coefficients,
        precrack_measurements=precrack_measurements,
        precrack_final_size=precrack_final_size,
        final_crack_measurements=final_crack_measurements,
        final_crack_size=final_crack_size,
        paris_law_C=paris_law_C,
        paris_law_m=paris_law_m,
        threshold_delta_K=threshold_delta_K,
        total_cycles=total_cycles,
        fcgr_results=fcgr_results,
        material='',
    )
