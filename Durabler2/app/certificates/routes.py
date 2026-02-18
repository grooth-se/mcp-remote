"""Certificate register routes."""
import os
from datetime import datetime, date
from pathlib import Path

from flask import (render_template, redirect, url_for, flash, request,
                   jsonify, current_app, send_file)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from . import certificates_bp
from .forms import CertificateForm, CertificateSearchForm, CertificateImportForm
from app.extensions import db
from app.models import AuditLog
from app.models.certificate import Certificate


@certificates_bp.route('/')
@login_required
def index():
    """List all certificates with filtering."""
    # Get filter parameters
    year_filter = request.args.get('year', '')
    search_term = request.args.get('search', '')

    # Build query
    query = Certificate.query

    if year_filter and year_filter != 'All':
        try:
            query = query.filter(Certificate.year == int(year_filter))
        except ValueError:
            pass

    if search_term:
        search_pattern = f'%{search_term}%'
        query = query.filter(
            db.or_(
                Certificate.test_order.ilike(search_pattern),
                Certificate.project_name.ilike(search_pattern),
                Certificate.customer.ilike(search_pattern),
                Certificate.customer_order.ilike(search_pattern),
                Certificate.product_sn.ilike(search_pattern),
                Certificate.test_article_sn.ilike(search_pattern),
                Certificate.customer_specimen_info.ilike(search_pattern),
                Certificate.material.ilike(search_pattern),
                Certificate.comment.ilike(search_pattern)
            )
        )

    # Order by year desc, cert_id desc
    certificates = query.order_by(
        Certificate.year.desc(),
        Certificate.cert_id.desc(),
        Certificate.revision.desc()
    ).limit(500).all()

    # Get years for filter dropdown
    years = Certificate.get_years_list()
    current_year = datetime.now().year
    if current_year not in years:
        years.insert(0, current_year)

    return render_template('certificates/index.html',
                           certificates=certificates,
                           years=years,
                           year_filter=year_filter,
                           search_term=search_term)


@certificates_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    """Create new certificate."""
    form = CertificateForm()

    if request.method == 'GET':
        # Pre-fill with defaults
        current_year = datetime.now().year
        next_id = Certificate.get_next_cert_id(current_year)
        form.year.data = current_year
        form.cert_id.data = next_id
        form.revision.data = 1
        form.cert_date.data = date.today()

    if form.validate_on_submit():
        # Check if certificate already exists
        existing = Certificate.query.filter_by(
            year=form.year.data,
            cert_id=form.cert_id.data,
            revision=form.revision.data
        ).first()

        if existing:
            flash(f'Certificate DUR-{form.year.data}-{form.cert_id.data} Rev.{form.revision.data} already exists!', 'danger')
            return render_template('certificates/edit.html', form=form, is_new=True)

        cert = Certificate(
            year=form.year.data,
            cert_id=form.cert_id.data,
            revision=form.revision.data,
            cert_date=form.cert_date.data,
            test_order=form.test_order.data,
            project_name=form.project_name.data,
            test_standard=form.test_standard.data,
            customer=form.customer.data,
            customer_order=form.customer_order.data,
            product=form.product.data,
            product_sn=form.product_sn.data,
            material=form.material.data,
            test_article_sn=form.test_article_sn.data,
            customer_specimen_info=form.customer_specimen_info.data,
            location_orientation=form.location_orientation.data,
            requirement=form.requirement.data,
            temperature=form.temperature.data,
            comment=form.comment.data,
            reported=form.reported.data,
            invoiced=form.invoiced.data,
            created_by_id=current_user.id
        )

        db.session.add(cert)

        # Audit log
        audit = AuditLog(
            user_id=current_user.id,
            action='CREATE',
            table_name='certificates',
            new_values={'certificate_number': cert.certificate_number},
            ip_address=request.remote_addr
        )
        db.session.add(audit)

        db.session.commit()

        flash(f'Certificate {cert.certificate_number} created successfully!', 'success')
        return redirect(url_for('certificates.view', cert_id=cert.id))

    return render_template('certificates/edit.html', form=form, is_new=True)


@certificates_bp.route('/<int:cert_id>')
@login_required
def view(cert_id):
    """View certificate details."""
    from app.models import ReportApproval, STATUS_COLORS, APPROVAL_STATUS_LABELS

    cert = Certificate.query.get_or_404(cert_id)

    # Get linked test records
    test_records = cert.test_records.all() if cert.test_records else []

    return render_template('certificates/view.html',
                           cert=cert,
                           test_records=test_records,
                           status_colors=STATUS_COLORS,
                           status_labels=APPROVAL_STATUS_LABELS)


@certificates_bp.route('/<int:cert_id>/edit', methods=['GET', 'POST'])
@login_required
def edit(cert_id):
    """Edit certificate."""
    cert = Certificate.query.get_or_404(cert_id)
    form = CertificateForm(obj=cert)

    if form.validate_on_submit():
        # Store old values for audit
        old_values = {
            'certificate_number': cert.certificate_number,
            'customer': cert.customer,
            'provningsorder': cert.test_order
        }

        # Update certificate
        cert.year = form.year.data
        cert.cert_id = form.cert_id.data
        cert.revision = form.revision.data
        cert.cert_date = form.cert_date.data
        cert.test_order = form.test_order.data
        cert.project_name = form.project_name.data
        cert.test_standard = form.test_standard.data
        cert.customer = form.customer.data
        cert.customer_order = form.customer_order.data
        cert.product = form.product.data
        cert.product_sn = form.product_sn.data
        cert.material = form.material.data
        cert.test_article_sn = form.test_article_sn.data
        cert.customer_specimen_info = form.customer_specimen_info.data
        cert.location_orientation = form.location_orientation.data
        cert.requirement = form.requirement.data
        cert.temperature = form.temperature.data
        cert.comment = form.comment.data
        cert.reported = form.reported.data
        cert.invoiced = form.invoiced.data

        # Audit log
        audit = AuditLog(
            user_id=current_user.id,
            action='UPDATE',
            table_name='certificates',
            record_id=cert.id,
            old_values=old_values,
            new_values={'certificate_number': cert.certificate_number},
            ip_address=request.remote_addr
        )
        db.session.add(audit)

        db.session.commit()

        flash(f'Certificate {cert.certificate_number} updated!', 'success')
        return redirect(url_for('certificates.view', cert_id=cert.id))

    return render_template('certificates/edit.html', form=form, cert=cert, is_new=False)


@certificates_bp.route('/<int:cert_id>/revision', methods=['POST'])
@login_required
def create_revision(cert_id):
    """Create new revision of certificate."""
    cert = Certificate.query.get_or_404(cert_id)

    # Create new certificate with incremented revision
    new_cert = Certificate(
        year=cert.year,
        cert_id=cert.cert_id,
        revision=cert.revision + 1,
        cert_date=date.today(),
        test_order=cert.test_order,
        project_name=cert.project_name,
        test_standard=cert.test_standard,
        customer=cert.customer,
        customer_order=cert.customer_order,
        product=cert.product,
        product_sn=cert.product_sn,
        material=cert.material,
        test_article_sn=cert.test_article_sn,
        customer_specimen_info=cert.customer_specimen_info,
        location_orientation=cert.location_orientation,
        requirement=cert.requirement,
        temperature=cert.temperature,
        comment=f"Revision of {cert.certificate_number_with_rev}",
        reported=False,
        invoiced=False,
        created_by_id=current_user.id
    )

    db.session.add(new_cert)

    # Audit log
    audit = AuditLog(
        user_id=current_user.id,
        action='CREATE',
        table_name='certificates',
        new_values={
            'certificate_number': new_cert.certificate_number_with_rev,
            'from_revision': cert.certificate_number_with_rev
        },
        ip_address=request.remote_addr
    )
    db.session.add(audit)

    db.session.commit()

    flash(f'Created revision {new_cert.certificate_number_with_rev}', 'success')
    return redirect(url_for('certificates.edit', cert_id=new_cert.id))


@certificates_bp.route('/<int:cert_id>/copy', methods=['POST'])
@login_required
def copy(cert_id):
    """Copy certificate to create a new one with next available ID."""
    cert = Certificate.query.get_or_404(cert_id)

    # Get next available ID for current year
    current_year = datetime.now().year
    next_id = Certificate.get_next_cert_id(current_year)

    # Create new certificate copying all data except cert number and status
    new_cert = Certificate(
        year=current_year,
        cert_id=next_id,
        revision=1,
        cert_date=date.today(),
        test_order=cert.test_order,
        project_name=cert.project_name,
        test_standard=cert.test_standard,
        customer=cert.customer,
        customer_order=cert.customer_order,
        product=cert.product,
        product_sn=cert.product_sn,
        material=cert.material,
        test_article_sn=cert.test_article_sn,
        customer_specimen_info=cert.customer_specimen_info,
        location_orientation=cert.location_orientation,
        requirement=cert.requirement,
        temperature=cert.temperature,
        comment=f"Copy of {cert.certificate_number_with_rev}",
        reported=False,
        invoiced=False,
        created_by_id=current_user.id
    )

    db.session.add(new_cert)

    # Audit log
    audit = AuditLog(
        user_id=current_user.id,
        action='COPY',
        table_name='certificates',
        new_values={
            'certificate_number': new_cert.certificate_number,
            'copied_from': cert.certificate_number_with_rev
        },
        ip_address=request.remote_addr
    )
    db.session.add(audit)

    db.session.commit()

    flash(f'Created {new_cert.certificate_number} (copy of {cert.certificate_number})', 'success')
    return redirect(url_for('certificates.edit', cert_id=new_cert.id))


@certificates_bp.route('/<int:cert_id>/delete', methods=['POST'])
@login_required
def delete(cert_id):
    """Delete certificate (admin only)."""
    if current_user.role != 'admin':
        flash('Only administrators can delete certificates.', 'danger')
        return redirect(url_for('certificates.index'))

    cert = Certificate.query.get_or_404(cert_id)

    # Check for linked test records
    if cert.test_records.count() > 0:
        flash(f'Cannot delete {cert.certificate_number} - has linked test records.', 'danger')
        return redirect(url_for('certificates.view', cert_id=cert_id))

    cert_num = cert.certificate_number_with_rev

    # Audit log
    audit = AuditLog(
        user_id=current_user.id,
        action='DELETE',
        table_name='certificates',
        record_id=cert_id,
        old_values={'certificate_number': cert_num},
        ip_address=request.remote_addr,
        reason=request.form.get('reason', 'User requested deletion')
    )
    db.session.add(audit)

    db.session.delete(cert)
    db.session.commit()

    flash(f'Certificate {cert_num} deleted.', 'success')
    return redirect(url_for('certificates.index'))


@certificates_bp.route('/search')
@login_required
def search():
    """Search certificates (for AJAX)."""
    term = request.args.get('q', '')

    if len(term) < 2:
        return jsonify([])

    # Search by certificate number or other fields
    pattern = f'%{term}%'

    # Try parsing as certificate number first
    year, cert_id, _ = Certificate.parse_certificate_number(term)

    if year and cert_id:
        certs = Certificate.query.filter_by(year=year, cert_id=cert_id)\
            .order_by(Certificate.revision.desc()).limit(10).all()
    else:
        certs = Certificate.query.filter(
            db.or_(
                Certificate.test_order.ilike(pattern),
                Certificate.customer.ilike(pattern),
                Certificate.product_sn.ilike(pattern),
                Certificate.test_article_sn.ilike(pattern)
            )
        ).order_by(Certificate.year.desc(), Certificate.cert_id.desc()).limit(10).all()

    return jsonify([{
        'id': c.id,
        'certificate_number': c.certificate_number,
        'certificate_number_with_rev': c.certificate_number_with_rev,
        'customer': c.customer or '',
        'provningsorder': c.test_order or ''
    } for c in certs])


@certificates_bp.route('/next-id')
@login_required
def next_id():
    """Get next certificate ID for year (AJAX)."""
    year = request.args.get('year', datetime.now().year, type=int)
    next_cert_id = Certificate.get_next_cert_id(year)
    return jsonify({'year': year, 'next_id': next_cert_id})


@certificates_bp.route('/<int:cert_id>/download-reports')
@login_required
def download_reports(cert_id):
    """Download all signed PDFs for a certificate as a ZIP file."""
    import zipfile
    import io

    cert = Certificate.query.get_or_404(cert_id)
    test_records = cert.test_records.all()

    # Find all published reports
    signed_files = []
    reports_folder = Path(current_app.config['REPORTS_FOLDER'])

    for test in test_records:
        if test.approval and test.approval.status == 'PUBLISHED' and test.approval.signed_pdf_path:
            pdf_path = reports_folder / test.approval.signed_pdf_path
            if pdf_path.exists():
                signed_files.append({
                    'path': pdf_path,
                    'name': f"{test.test_id}_{test.test_method}.pdf"
                })

    if not signed_files:
        flash('No signed PDFs available for this certificate.', 'warning')
        return redirect(url_for('certificates.view', cert_id=cert_id))

    # Create ZIP file in memory
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for file_info in signed_files:
            zip_file.write(file_info['path'], file_info['name'])

    zip_buffer.seek(0)

    # Generate filename
    zip_filename = f"{cert.certificate_number.replace('-', '_')}_signed_reports.zip"

    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name=zip_filename,
        mimetype='application/zip'
    )


@certificates_bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_excel():
    """Import certificates from Excel file."""
    if current_user.role != 'admin':
        flash('Only administrators can import certificates.', 'danger')
        return redirect(url_for('certificates.index'))

    form = CertificateImportForm()

    if form.validate_on_submit():
        try:
            from openpyxl import load_workbook

            file = form.excel_file.data
            filename = secure_filename(file.filename)
            filepath = Path(current_app.config['UPLOAD_FOLDER']) / filename
            file.save(filepath)

            # Load workbook
            wb = load_workbook(filepath)
            ws = wb.active

            # Get headers from first row
            headers = [cell.value for cell in ws[1]]

            # Map column names to indices
            col_map = {}
            header_mapping = {
                'Year': 'year',
                'ID': 'cert_id',
                'Revision No': 'revision',
                'Date': 'cert_date',
                'Product': 'product',
                'Product S/N': 'product_sn',
                'Provningsorder': 'provningsorder',
                'Project name': 'project_name',
                'Test standard': 'test_standard',
                'Material/ HT': 'material',
                'Test article SN': 'test_article_sn',
                'Customer Specimen Info': 'customer_specimen_info',
                'Location/orient.:': 'location_orientation',
                'Requirement': 'requirement',
                'Temperature': 'temperature',
                'Customer': 'customer',
                'Customer order': 'customer_order',
                'Comment': 'comment',
                'Reported': 'reported',
                'Invoiced': 'invoiced'
            }

            for idx, header in enumerate(headers):
                if header in header_mapping:
                    col_map[header_mapping[header]] = idx

            # Process rows
            imported = 0
            skipped = 0
            errors = []

            for row_num in range(2, ws.max_row + 1):
                row_data = [cell.value for cell in ws[row_num]]

                # Skip empty rows
                if not row_data[col_map.get('year', 0)]:
                    continue

                try:
                    year = int(row_data[col_map.get('year', 0)])
                    cert_id = int(row_data[col_map.get('cert_id', 1)])
                    revision = int(row_data[col_map.get('revision', 3)] or 1)

                    # Check if exists
                    existing = Certificate.query.filter_by(
                        year=year, cert_id=cert_id, revision=revision
                    ).first()

                    if existing:
                        if form.skip_existing.data:
                            skipped += 1
                            continue
                        else:
                            # Update existing
                            cert = existing
                    else:
                        cert = Certificate(
                            year=year,
                            cert_id=cert_id,
                            revision=revision,
                            created_by_id=current_user.id
                        )
                        db.session.add(cert)

                    # Set other fields
                    cert_date = row_data[col_map.get('cert_date')] if col_map.get('cert_date') is not None else None
                    if cert_date:
                        if isinstance(cert_date, datetime):
                            cert.cert_date = cert_date.date()
                        elif isinstance(cert_date, date):
                            cert.cert_date = cert_date

                    # Helper to get string value from column
                    def get_str(field):
                        idx = col_map.get(field)
                        if idx is not None and idx < len(row_data):
                            return str(row_data[idx] or '')
                        return ''

                    cert.product = get_str('product')
                    cert.product_sn = get_str('product_sn')
                    cert.test_order = get_str('provningsorder')
                    cert.project_name = get_str('project_name')
                    cert.test_standard = get_str('test_standard')
                    cert.material = get_str('material')
                    cert.test_article_sn = get_str('test_article_sn')
                    cert.customer_specimen_info = get_str('customer_specimen_info')
                    cert.location_orientation = get_str('location_orientation')
                    cert.requirement = get_str('requirement')
                    cert.temperature = get_str('temperature')
                    cert.customer = get_str('customer')
                    cert.customer_order = get_str('customer_order')
                    cert.comment = get_str('comment')

                    # Boolean fields - 'X' or 'x' means True
                    reported_idx = col_map.get('reported')
                    invoiced_idx = col_map.get('invoiced')
                    reported = row_data[reported_idx] if reported_idx is not None and reported_idx < len(row_data) else None
                    invoiced = row_data[invoiced_idx] if invoiced_idx is not None and invoiced_idx < len(row_data) else None
                    cert.reported = str(reported).upper() == 'X' if reported else False
                    cert.invoiced = str(invoiced).upper() == 'X' if invoiced else False

                    imported += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")

            db.session.commit()

            # Audit log
            audit = AuditLog(
                user_id=current_user.id,
                action='IMPORT',
                table_name='certificates',
                new_values={
                    'filename': filename,
                    'imported': imported,
                    'skipped': skipped,
                    'errors': len(errors)
                },
                ip_address=request.remote_addr
            )
            db.session.add(audit)
            db.session.commit()

            # Clean up uploaded file
            os.remove(filepath)

            flash(f'Import complete: {imported} imported, {skipped} skipped, {len(errors)} errors', 'success')

            if errors:
                for err in errors[:5]:  # Show first 5 errors
                    flash(err, 'warning')
                if len(errors) > 5:
                    flash(f'...and {len(errors) - 5} more errors', 'warning')

            return redirect(url_for('certificates.index'))

        except Exception as e:
            import traceback
            traceback.print_exc()
            flash(f'Import error: {str(e)}', 'danger')

    return render_template('certificates/import.html', form=form)
