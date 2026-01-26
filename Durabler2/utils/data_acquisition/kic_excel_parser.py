"""
KIC Excel file parser for MTS TestSuite Analysis Reports.

Imports specimen dimensions, material properties, precrack measurements,
and KIC results from Excel files exported by MTS TestSuite.
"""

from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, List, Dict
import numpy as np


@dataclass
class KICUserInputs:
    """
    Container for KIC user input data imported from Excel.

    Parameters
    ----------
    specimen_id : str
        Specimen identifier (Test Run name)
    specimen_type : str
        'SE(B)' or 'C(T)'
    W : float
        Specimen width (mm)
    B : float
        Specimen thickness (mm)
    B_n : float
        Net thickness at side grooves (mm)
    a_0 : float
        Initial notch length (mm)
    S : float
        Span for SE(B) specimens (mm)
    yield_strength : float
        0.2% yield strength (MPa)
    ultimate_strength : float
        Ultimate tensile strength (MPa)
    youngs_modulus : float
        Young's modulus (GPa)
    poissons_ratio : float
        Poisson's ratio
    test_temperature : float
        Test temperature (C)
    precrack_measurements : List[float]
        Precrack crack size measurements (mm)
    precrack_final_size : float
        Final precrack size from compliance (mm)
    k_calibration_coefficients : List[float]
        K calibration coefficients from standard
    kic_results : Dict[str, any]
        KIC results from MTS analysis
    """
    specimen_id: str
    specimen_type: str
    W: float
    B: float
    B_n: float
    a_0: float
    S: float
    yield_strength: float
    ultimate_strength: float
    youngs_modulus: float
    poissons_ratio: float = 0.3
    test_temperature: float = 23.0
    precrack_measurements: List[float] = field(default_factory=list)
    precrack_final_size: float = 0.0
    k_calibration_coefficients: List[float] = field(default_factory=list)
    kic_results: Dict[str, any] = field(default_factory=dict)
    material: str = ""

    @property
    def crack_length_average(self) -> float:
        """Calculate average crack length from measurements."""
        if not self.precrack_measurements:
            return self.a_0

        # If we have 5 measurements (typical for KIC per E399)
        if len(self.precrack_measurements) == 5:
            a = self.precrack_measurements
            # E399 formula: a = (a1 + a2 + a3 + a4 + a5) / 5
            # But a1 and a5 are quarter points, weighted by 0.5
            # Actually E399 uses: a = (0.5*a1 + a2 + a3 + a4 + 0.5*a5) / 4
            return (0.5 * a[0] + a[1] + a[2] + a[3] + 0.5 * a[4]) / 4

        # For 9 measurements (like E1820/E1290)
        if len(self.precrack_measurements) == 9:
            a = self.precrack_measurements
            return (0.5 * a[0] + sum(a[1:8]) + 0.5 * a[8]) / 8

        return np.mean(self.precrack_measurements)

    @property
    def a_W_ratio(self) -> float:
        """Calculate a/W ratio."""
        if self.W == 0:
            return 0.0
        a = self.crack_length_average if self.precrack_measurements else self.a_0
        return a / self.W


def parse_kic_excel(filepath: Path) -> KICUserInputs:
    """
    Parse KIC user input data from MTS Excel Analysis Report.

    The Excel file contains specimen dimensions, material properties,
    precrack measurements, and KIC analysis results.

    Parameters
    ----------
    filepath : Path
        Path to Excel file (.xlsx)

    Returns
    -------
    KICUserInputs
        Parsed user input data

    Raises
    ------
    FileNotFoundError
        If file does not exist
    ValueError
        If required data cannot be found
    """
    from openpyxl import load_workbook

    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    try:
        wb = load_workbook(filepath, data_only=True)
        # Try 'Data' sheet first, then first sheet
        if 'Data' in wb.sheetnames:
            ws = wb['Data']
        else:
            ws = wb.active
    except Exception as e:
        raise ValueError(f"Failed to read Excel file: {e}")

    # Build a dictionary of all label-value pairs
    # MTS format: label in column C (index 3), value in column I (index 9)
    data_dict = {}
    for row in ws.iter_rows(min_row=1, max_row=200, values_only=True):
        if row and len(row) >= 9:
            label = row[2] if len(row) > 2 else None  # Column C
            value = row[8] if len(row) > 8 else None  # Column I
            if label and value is not None:
                data_dict[str(label).strip()] = value

    def parse_value(val_str, default=0.0):
        """Extract numeric value from string like '83.1000 mm' or '0.6900 kN/mm²'."""
        if val_str is None:
            return default
        val_str = str(val_str)
        # Remove units and extract number
        parts = val_str.split()
        if parts:
            try:
                return float(parts[0])
            except ValueError:
                return default
        return default

    def get_value(keys, default=0.0):
        """Get value from data_dict using multiple possible keys."""
        for key in keys:
            if key in data_dict:
                return parse_value(data_dict[key], default)
        return default

    def get_string(keys, default=''):
        """Get string value from data_dict."""
        for key in keys:
            if key in data_dict:
                val = data_dict[key]
                if val:
                    return str(val)
        return default

    # Extract specimen info
    specimen_id = get_string(['Name of the Test Run', 'Specimen Name'], 'Unknown')
    geometry_type = get_string(['Geometry Type'], 'SE(B)')

    # Determine specimen type from geometry string
    geom_lower = geometry_type.lower()
    if 'c(t)' in geom_lower or 'ct' in geom_lower or 'compact' in geom_lower:
        specimen_type = 'C(T)'
    else:
        specimen_type = 'SE(B)'

    # Specimen dimensions
    W = get_value(['Width (W)'], 25.0)
    B = get_value(['Thickness (B)'], 12.5)
    B_n = get_value(['Net Thickness (Bn)'], B)
    a_0 = get_value(['Notch Length (a0)'], 12.5)
    S = get_value(['Span (S)'], W * 4 if specimen_type == 'SE(B)' else 0)

    # Material properties (MTS uses kN/mm² for stress, convert to MPa)
    E_raw = get_value(['Elastic Modulus (E)'], 210.0)
    youngs_modulus = E_raw  # kN/mm² = GPa

    yield_raw = get_value(['Yield Strength'], 0.5)
    yield_strength = yield_raw * 1000  # kN/mm² to MPa

    uts_raw = get_value(['Ultimate Tensile Strength'], 0.6)
    ultimate_strength = uts_raw * 1000  # kN/mm² to MPa

    poissons_ratio = get_value(["Poisson's Ratio"], 0.3)
    test_temperature = get_value(['Valid at Temperature'], 23.0)

    # Precrack final size from compliance
    precrack_final_size = get_value(['Precrack Final Crack Size'], a_0)

    # K calibration coefficients
    k_calibration_coefficients = []
    for i in range(5):
        coef = get_value([f'K Calibration Coefficient {i}'], 0.0)
        k_calibration_coefficients.append(coef)

    # Precrack measurements (5 points for E399)
    precrack_measurements = []
    for i in range(1, 10):
        val = get_value([f'Precrack Crack {i}'], 0.0)
        if val > 0:
            precrack_measurements.append(val)

    # KIC results from MTS analysis
    kic_results = {
        'KQ': get_value(['KQ'], 0.0),
        'PQ': get_value(['PQ'], 0.0),
        'P5': get_value(['P5'], 0.0),
        'P_max': get_value(['P Maximum'], 0.0),
        'type': get_string(['Type'], ''),
        'r_squared': get_value(['r^2 Correlation Coefficient'], 0.0),
        'strength_ratio': get_value(['Strength Ratio (Rs)'], 0.0),
        # Validity checks
        'pmax_pq_valid': get_string(['Is PMax / PQ <= 1.1'], ''),
        'pmax_pq_ratio': get_value(['Pmax / PQ'], 0.0),
        'plane_strain_valid': get_string(['Is 2.5 (KQ/SYS)^2 < W - a ?'], ''),
        'aw_ratio_valid': get_string(['Is a/W between 0.45 and 0.55?'], ''),
        'a_W': get_value(['a/W'], 0.0),
        'avg_crack_size': get_value(['Average Crack Size (a)'], 0.0),
        'ligament': get_value(['Ligament (W - a)'], 0.0),
        'specimen_thickness_req': get_value(['Specimen Thickness (2.5 (KQ/SYS)^2)'], 0.0),
        'loading_rate': get_value(['Loading Rate'], 0.0),
        'loading_rate_valid': get_string(['Is loading rate between valid limit'], ''),
    }

    return KICUserInputs(
        specimen_id=str(specimen_id),
        specimen_type=specimen_type,
        W=W,
        B=B,
        B_n=B_n if B_n else B,
        a_0=a_0,
        S=S,
        yield_strength=yield_strength,
        ultimate_strength=ultimate_strength,
        youngs_modulus=youngs_modulus,
        poissons_ratio=poissons_ratio,
        test_temperature=test_temperature,
        precrack_measurements=precrack_measurements,
        precrack_final_size=precrack_final_size,
        k_calibration_coefficients=k_calibration_coefficients,
        kic_results=kic_results,
        material='',
    )
