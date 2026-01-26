"""
CTOD Excel file parser for user input data.

Imports specimen dimensions, material properties, and crack measurements
from Excel files exported by MTS TestSuite or user-created spreadsheets.
"""

from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, List, Dict
import numpy as np


@dataclass
class CTODUserInputs:
    """
    Container for CTOD user input data imported from Excel.

    Parameters
    ----------
    specimen_id : str
        Specimen identifier
    specimen_type : str
        'SE(B)' or 'C(T)'
    W : float
        Specimen width (mm)
    B : float
        Specimen thickness (mm)
    B_n : float
        Net thickness at side grooves (mm)
    a_0 : float
        Initial crack length (mm)
    S : float
        Span for SE(B) specimens (mm)
    yield_strength : float
        0.2% yield strength (MPa)
    ultimate_strength : float
        Ultimate tensile strength (MPa)
    youngs_modulus : float
        Young's modulus (GPa)
    test_temperature : float
        Test temperature (°C)
    precrack_measurements : List[float]
        9-point pre-crack length measurements (mm)
    final_crack_measurements : List[float]
        9-point final crack length measurements (mm)
    material : str
        Material designation
    compliance_coefficients : List[float]
        Compliance coefficients C0-C5 for crack length calculation
    ctod_results : Dict[str, float]
        CTOD results from MTS (CTODc, CTODu, P_max, etc.)
    """
    specimen_id: str
    specimen_type: str
    W: float
    B: float
    B_n: float
    a_0: float
    S: float
    yield_strength: float
    ultimate_strength: float
    youngs_modulus: float
    test_temperature: float
    precrack_measurements: List[float]
    final_crack_measurements: List[float] = field(default_factory=list)
    material: str = ""
    compliance_coefficients: List[float] = field(default_factory=list)
    ctod_results: Dict[str, float] = field(default_factory=dict)
    poissons_ratio: float = 0.3

    # Alias for backward compatibility
    @property
    def crack_measurements(self) -> List[float]:
        """Alias for precrack_measurements for backward compatibility."""
        return self.precrack_measurements

    @property
    def crack_length_average(self) -> float:
        """
        Calculate average crack length from 9-point measurements.

        Per ASTM E1290, the average is calculated as:
        a_avg = (0.5*a1 + a2 + a3 + ... + a8 + 0.5*a9) / 8

        Where a1 and a9 are surface measurements (weighted half) and a2-a8 are interior.

        Returns
        -------
        float
            Average crack length (mm)
        """
        if len(self.precrack_measurements) != 9:
            return np.mean(self.precrack_measurements) if self.precrack_measurements else 0.0

        a = self.precrack_measurements
        # Surface points (a1, a9) weighted 0.5, interior points full weight
        return (0.5 * a[0] + sum(a[1:8]) + 0.5 * a[8]) / 8

    @property
    def final_crack_length_average(self) -> float:
        """Calculate average final crack length from 9-point measurements."""
        if len(self.final_crack_measurements) != 9:
            return np.mean(self.final_crack_measurements) if self.final_crack_measurements else 0.0

        a = self.final_crack_measurements
        return (0.5 * a[0] + sum(a[1:8]) + 0.5 * a[8]) / 8


def parse_ctod_excel(filepath: Path) -> CTODUserInputs:
    """
    Parse CTOD user input data from Excel file.

    The Excel file should contain specimen dimensions, material properties,
    and crack measurements. This parser handles MTS CTOD Analysis Report format.

    Parameters
    ----------
    filepath : Path
        Path to Excel file (.xlsx or .xls)

    Returns
    -------
    CTODUserInputs
        Parsed user input data

    Raises
    ------
    FileNotFoundError
        If file does not exist
    ValueError
        If required data cannot be found
    """
    from openpyxl import load_workbook

    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    try:
        wb = load_workbook(filepath)
        # Try 'Data' sheet first, then first sheet
        if 'Data' in wb.sheetnames:
            ws = wb['Data']
        else:
            ws = wb.active
    except Exception as e:
        raise ValueError(f"Failed to read Excel file: {e}")

    # Build a dictionary of all label-value pairs
    data_dict = {}
    for row in ws.iter_rows(values_only=True):
        if row and len(row) >= 8:
            # MTS format: label in column B (index 1), value in column H (index 7)
            label = row[1] if len(row) > 1 else None
            value = row[7] if len(row) > 7 else None
            if label and value:
                data_dict[str(label).strip()] = value

    def parse_value(val_str, default=0.0):
        """Extract numeric value from string like '83.1000 mm' or '0.6900 kN/mm²'."""
        if val_str is None:
            return default
        val_str = str(val_str)
        # Remove units and extract number
        parts = val_str.split()
        if parts:
            try:
                return float(parts[0])
            except ValueError:
                return default
        return default

    def get_value(keys, default=0.0):
        """Get value from data_dict using multiple possible keys."""
        for key in keys:
            if key in data_dict:
                return parse_value(data_dict[key], default)
        return default

    def get_string(keys, default=''):
        """Get string value from data_dict."""
        for key in keys:
            if key in data_dict:
                val = data_dict[key]
                if val:
                    return str(val).split()[0] if ' ' in str(val) else str(val)
        return default

    # Extract specimen info
    specimen_id = get_string(['Name of the Test Run', 'Specimen Name'], 'Unknown')
    geometry_type = get_string(['Geometry Type'], 'SE(B)')
    specimen_type = 'C(T)' if 'c(t)' in geometry_type.lower() else 'SE(B)'

    # Specimen dimensions
    W = get_value(['Width (W)'], 25.0)
    B = get_value(['Thickness (B)'], 12.5)
    B_n = get_value(['Net Thickness (Bn)'], B)
    a_0 = get_value(['Notch Length (a0)'], 12.5)
    S = get_value(['Span (S)'], W * 4)

    # Material properties (MTS uses kN/mm² for stress, convert to MPa)
    E_raw = get_value(['Elastic Modulus (E)'], 210.0)
    # E is in kN/mm² = GPa, keep as GPa
    youngs_modulus = E_raw

    yield_raw = get_value(['Yield Strength'], 0.5)
    # Yield in kN/mm² = GPa, convert to MPa (* 1000)
    yield_strength = yield_raw * 1000

    uts_raw = get_value(['Ultimate Tensile Strength'], 0.6)
    ultimate_strength = uts_raw * 1000

    poissons_ratio = get_value(["Poisson's Ratio"], 0.3)
    test_temperature = get_value(['Valid at Temperature'], 23.0)

    # Compliance coefficients
    compliance_coefficients = []
    for i in range(6):
        coef = get_value([f'Compliance Coefficient C{i}'], 0.0)
        compliance_coefficients.append(coef)

    # Precrack measurements (9-point)
    precrack_measurements = []
    for i in range(1, 10):
        val = get_value([f'Precrack Crack {i}'], 0.0)
        if val > 0:
            precrack_measurements.append(val)

    # Final crack measurements (9-point)
    final_crack_measurements = []
    for i in range(1, 10):
        val = get_value([f'Final Crack {i}'], 0.0)
        if val > 0:
            final_crack_measurements.append(val)

    # CTOD results from MTS
    ctod_results = {
        'CTODc': get_value(['CTODc'], 0.0),
        'CTODu': get_value(['CTODu'], 0.0),
        'CTOD_max': get_value(['CTOD Maximum'], 0.0),
        'P_max': get_value(['P Maximum'], 0.0),
        'Pu': get_value(['Pu'], 0.0),
        'Pf': get_value(['Pf'], 0.0),
        'a_W': get_value(['a/W'], 0.0),
        'avg_precrack': get_value(['Average Precrack Size'], 0.0),
        'avg_final_crack': get_value(['Average Final Crack Size'], 0.0),
        'precrack_cycles': get_value(['Precrack Cycles Completed'], 0),
        'precrack_final_K': get_value(['Precrack Final K'], 0.0),
    }

    # If no precrack measurements found, use a_0
    if not precrack_measurements or len(precrack_measurements) < 9:
        precrack_measurements = [a_0] * 9

    return CTODUserInputs(
        specimen_id=str(specimen_id),
        specimen_type=specimen_type,
        W=W,
        B=B,
        B_n=B_n if B_n else B,
        a_0=a_0,
        S=S,
        yield_strength=yield_strength,
        ultimate_strength=ultimate_strength,
        youngs_modulus=youngs_modulus,
        test_temperature=test_temperature,
        precrack_measurements=precrack_measurements,
        final_crack_measurements=final_crack_measurements,
        material='',
        compliance_coefficients=compliance_coefficients,
        ctod_results=ctod_results,
        poissons_ratio=poissons_ratio
    )
