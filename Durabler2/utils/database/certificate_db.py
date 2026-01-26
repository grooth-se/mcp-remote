"""
Certificate Register Database Module.

SQLite database for managing test certificate numbers and associated test information.
Certificate format: DUR-XXXX-YYYY where XXXX=year, YYYY=sequential number (starts 1001/year).
"""

import sqlite3
from pathlib import Path
from datetime import datetime, date
from typing import Optional, List, Dict, Any
from dataclasses import dataclass, asdict


# Default database path
DEFAULT_DB_PATH = Path(__file__).parent.parent.parent / "data" / "certificate_register.db"


@dataclass
class Certificate:
    """Certificate record data class."""
    id: Optional[int] = None  # Database row ID
    year: int = 0
    cert_id: int = 0  # Sequential number (1001, 1002, etc.)
    revision: int = 1
    cert_date: Optional[str] = None  # ISO format date string
    product: str = ""
    product_sn: str = ""
    test_project: str = ""
    project_name: str = ""
    test_standard: str = ""
    material: str = ""
    specimen_id: str = ""
    location_orientation: str = ""
    temperature: str = ""
    customer: str = ""
    customer_order: str = ""
    comment: str = ""
    reported: bool = False
    invoiced: bool = False
    created_at: Optional[str] = None
    updated_at: Optional[str] = None

    @property
    def certificate_number(self) -> str:
        """Generate full certificate number string."""
        return f"DUR-{self.year}-{self.cert_id}"

    @property
    def certificate_number_with_rev(self) -> str:
        """Generate certificate number with revision."""
        return f"DUR-{self.year}-{self.cert_id} Rev.{self.revision}"

    @classmethod
    def from_row(cls, row: sqlite3.Row) -> 'Certificate':
        """Create Certificate from database row."""
        return cls(
            id=row['id'],
            year=row['year'],
            cert_id=row['cert_id'],
            revision=row['revision'],
            cert_date=row['cert_date'],
            product=row['product'] or "",
            product_sn=row['product_sn'] or "",
            test_project=row['test_project'] or "",
            project_name=row['project_name'] or "",
            test_standard=row['test_standard'] or "",
            material=row['material'] or "",
            specimen_id=row['specimen_id'] or "",
            location_orientation=row['location_orientation'] or "",
            temperature=row['temperature'] or "",
            customer=row['customer'] or "",
            customer_order=row['customer_order'] or "",
            comment=row['comment'] or "",
            reported=bool(row['reported']),
            invoiced=bool(row['invoiced']),
            created_at=row['created_at'],
            updated_at=row['updated_at']
        )


class CertificateDatabase:
    """
    SQLite database manager for certificate register.

    Handles CRUD operations for test certificates.
    """

    def __init__(self, db_path: Optional[Path] = None):
        """
        Initialize database connection.

        Parameters
        ----------
        db_path : Path, optional
            Path to SQLite database file. Uses default if not specified.
        """
        self.db_path = db_path or DEFAULT_DB_PATH
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._init_database()

    def _get_connection(self) -> sqlite3.Connection:
        """Get database connection with row factory."""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_database(self):
        """Initialize database schema."""
        conn = self._get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS certificates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                year INTEGER NOT NULL,
                cert_id INTEGER NOT NULL,
                revision INTEGER DEFAULT 1,
                cert_date TEXT,
                product TEXT,
                product_sn TEXT,
                test_project TEXT,
                project_name TEXT,
                test_standard TEXT,
                material TEXT,
                specimen_id TEXT,
                location_orientation TEXT,
                temperature TEXT,
                customer TEXT,
                customer_order TEXT,
                comment TEXT,
                reported INTEGER DEFAULT 0,
                invoiced INTEGER DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(year, cert_id, revision)
            )
        """)

        # Create index for fast lookups
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_cert_year_id
            ON certificates(year, cert_id)
        """)

        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_cert_number
            ON certificates(year, cert_id, revision)
        """)

        conn.commit()
        conn.close()

    def get_next_cert_id(self, year: Optional[int] = None) -> int:
        """
        Get next available certificate ID for the given year.

        Parameters
        ----------
        year : int, optional
            Year for certificate. Defaults to current year.

        Returns
        -------
        int
            Next available certificate ID (starts at 1001)
        """
        if year is None:
            year = datetime.now().year

        conn = self._get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT MAX(cert_id) FROM certificates WHERE year = ?
        """, (year,))

        result = cursor.fetchone()[0]
        conn.close()

        if result is None:
            return 1001  # Start at 1001 for new year
        return result + 1

    def create_certificate(self, cert: Certificate) -> int:
        """
        Create new certificate record.

        Parameters
        ----------
        cert : Certificate
            Certificate data to insert

        Returns
        -------
        int
            New record ID
        """
        conn = self._get_connection()
        cursor = conn.cursor()

        now = datetime.now().isoformat()

        cursor.execute("""
            INSERT INTO certificates (
                year, cert_id, revision, cert_date, product, product_sn,
                test_project, project_name, test_standard, material,
                specimen_id, location_orientation, temperature, customer,
                customer_order, comment, reported, invoiced, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            cert.year, cert.cert_id, cert.revision, cert.cert_date,
            cert.product, cert.product_sn, cert.test_project, cert.project_name,
            cert.test_standard, cert.material, cert.specimen_id,
            cert.location_orientation, cert.temperature, cert.customer,
            cert.customer_order, cert.comment, int(cert.reported),
            int(cert.invoiced), now, now
        ))

        record_id = cursor.lastrowid
        conn.commit()
        conn.close()

        return record_id

    def update_certificate(self, cert: Certificate) -> bool:
        """
        Update existing certificate record.

        Parameters
        ----------
        cert : Certificate
            Certificate data with id set

        Returns
        -------
        bool
            True if updated successfully
        """
        if cert.id is None:
            return False

        conn = self._get_connection()
        cursor = conn.cursor()

        now = datetime.now().isoformat()

        cursor.execute("""
            UPDATE certificates SET
                year = ?, cert_id = ?, revision = ?, cert_date = ?,
                product = ?, product_sn = ?, test_project = ?, project_name = ?,
                test_standard = ?, material = ?, specimen_id = ?,
                location_orientation = ?, temperature = ?, customer = ?,
                customer_order = ?, comment = ?, reported = ?, invoiced = ?,
                updated_at = ?
            WHERE id = ?
        """, (
            cert.year, cert.cert_id, cert.revision, cert.cert_date,
            cert.product, cert.product_sn, cert.test_project, cert.project_name,
            cert.test_standard, cert.material, cert.specimen_id,
            cert.location_orientation, cert.temperature, cert.customer,
            cert.customer_order, cert.comment, int(cert.reported),
            int(cert.invoiced), now, cert.id
        ))

        success = cursor.rowcount > 0
        conn.commit()
        conn.close()

        return success

    def delete_certificate(self, record_id: int) -> bool:
        """
        Delete certificate record.

        Parameters
        ----------
        record_id : int
            Database record ID

        Returns
        -------
        bool
            True if deleted successfully
        """
        conn = self._get_connection()
        cursor = conn.cursor()

        cursor.execute("DELETE FROM certificates WHERE id = ?", (record_id,))

        success = cursor.rowcount > 0
        conn.commit()
        conn.close()

        return success

    def get_certificate_by_id(self, record_id: int) -> Optional[Certificate]:
        """Get certificate by database ID."""
        conn = self._get_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM certificates WHERE id = ?", (record_id,))
        row = cursor.fetchone()
        conn.close()

        if row:
            return Certificate.from_row(row)
        return None

    def get_certificate_by_number(
        self,
        year: int,
        cert_id: int,
        revision: Optional[int] = None
    ) -> Optional[Certificate]:
        """
        Get certificate by certificate number.

        Parameters
        ----------
        year : int
            Certificate year
        cert_id : int
            Certificate sequential ID
        revision : int, optional
            Specific revision. If None, returns latest revision.

        Returns
        -------
        Certificate or None
        """
        conn = self._get_connection()
        cursor = conn.cursor()

        if revision is not None:
            cursor.execute("""
                SELECT * FROM certificates
                WHERE year = ? AND cert_id = ? AND revision = ?
            """, (year, cert_id, revision))
        else:
            cursor.execute("""
                SELECT * FROM certificates
                WHERE year = ? AND cert_id = ?
                ORDER BY revision DESC LIMIT 1
            """, (year, cert_id))

        row = cursor.fetchone()
        conn.close()

        if row:
            return Certificate.from_row(row)
        return None

    def parse_certificate_number(self, cert_num: str) -> tuple:
        """
        Parse certificate number string.

        Parameters
        ----------
        cert_num : str
            Certificate number like "DUR-2024-1001" or "DUR-2024-1001 Rev.2"

        Returns
        -------
        tuple
            (year, cert_id, revision) or (None, None, None) if invalid
        """
        import re

        # Match DUR-YYYY-NNNN [Rev.R]
        match = re.match(r'DUR-(\d{4})-(\d+)(?:\s*Rev\.?(\d+))?', cert_num, re.IGNORECASE)
        if match:
            year = int(match.group(1))
            cert_id = int(match.group(2))
            revision = int(match.group(3)) if match.group(3) else None
            return year, cert_id, revision
        return None, None, None

    def get_certificate_by_string(self, cert_num: str) -> Optional[Certificate]:
        """Get certificate by certificate number string."""
        year, cert_id, revision = self.parse_certificate_number(cert_num)
        if year and cert_id:
            return self.get_certificate_by_number(year, cert_id, revision)
        return None

    def get_all_certificates(
        self,
        year: Optional[int] = None,
        limit: int = 1000,
        offset: int = 0
    ) -> List[Certificate]:
        """
        Get all certificates, optionally filtered by year.

        Parameters
        ----------
        year : int, optional
            Filter by year
        limit : int
            Maximum records to return
        offset : int
            Skip first N records

        Returns
        -------
        List[Certificate]
        """
        conn = self._get_connection()
        cursor = conn.cursor()

        if year:
            cursor.execute("""
                SELECT * FROM certificates
                WHERE year = ?
                ORDER BY year DESC, cert_id DESC, revision DESC
                LIMIT ? OFFSET ?
            """, (year, limit, offset))
        else:
            cursor.execute("""
                SELECT * FROM certificates
                ORDER BY year DESC, cert_id DESC, revision DESC
                LIMIT ? OFFSET ?
            """, (limit, offset))

        rows = cursor.fetchall()
        conn.close()

        return [Certificate.from_row(row) for row in rows]

    def get_certificate_numbers_list(self, year: Optional[int] = None) -> List[str]:
        """
        Get list of certificate numbers for dropdown/selection.

        Parameters
        ----------
        year : int, optional
            Filter by year

        Returns
        -------
        List[str]
            List of certificate numbers like "DUR-2024-1001"
        """
        conn = self._get_connection()
        cursor = conn.cursor()

        if year:
            cursor.execute("""
                SELECT DISTINCT year, cert_id
                FROM certificates
                WHERE year = ?
                ORDER BY year DESC, cert_id DESC
            """, (year,))
        else:
            cursor.execute("""
                SELECT DISTINCT year, cert_id
                FROM certificates
                ORDER BY year DESC, cert_id DESC
            """)

        rows = cursor.fetchall()
        conn.close()

        return [f"DUR-{row['year']}-{row['cert_id']}" for row in rows]

    def search_certificates(self, search_term: str, limit: int = 100) -> List[Certificate]:
        """
        Search certificates by various fields.

        Parameters
        ----------
        search_term : str
            Search term to match against multiple fields
        limit : int
            Maximum results

        Returns
        -------
        List[Certificate]
        """
        conn = self._get_connection()
        cursor = conn.cursor()

        search_pattern = f"%{search_term}%"

        cursor.execute("""
            SELECT * FROM certificates
            WHERE product_sn LIKE ?
               OR test_project LIKE ?
               OR project_name LIKE ?
               OR customer LIKE ?
               OR customer_order LIKE ?
               OR specimen_id LIKE ?
               OR material LIKE ?
               OR comment LIKE ?
            ORDER BY year DESC, cert_id DESC, revision DESC
            LIMIT ?
        """, (search_pattern,) * 8 + (limit,))

        rows = cursor.fetchall()
        conn.close()

        return [Certificate.from_row(row) for row in rows]

    def get_years_list(self) -> List[int]:
        """Get list of years that have certificates."""
        conn = self._get_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT DISTINCT year FROM certificates ORDER BY year DESC")
        rows = cursor.fetchall()
        conn.close()

        return [row['year'] for row in rows]

    def get_record_count(self, year: Optional[int] = None) -> int:
        """Get total number of certificate records."""
        conn = self._get_connection()
        cursor = conn.cursor()

        if year:
            cursor.execute("SELECT COUNT(*) FROM certificates WHERE year = ?", (year,))
        else:
            cursor.execute("SELECT COUNT(*) FROM certificates")

        count = cursor.fetchone()[0]
        conn.close()

        return count

    def import_from_excel(self, excel_path: Path) -> int:
        """
        Import certificates from Excel file.

        Parameters
        ----------
        excel_path : Path
            Path to Excel file

        Returns
        -------
        int
            Number of records imported
        """
        from openpyxl import load_workbook

        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active

        imported = 0

        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not row[0]:
                continue

            year = int(row[0]) if row[0] else datetime.now().year
            cert_id = int(row[1]) if row[1] else self.get_next_cert_id(year)
            revision = int(row[3]) if row[3] else 1

            # Check if already exists
            existing = self.get_certificate_by_number(year, cert_id, revision)
            if existing:
                continue

            # Parse date
            cert_date = None
            if row[4]:
                if isinstance(row[4], datetime):
                    cert_date = row[4].strftime("%Y-%m-%d")
                elif isinstance(row[4], str):
                    cert_date = row[4]

            cert = Certificate(
                year=year,
                cert_id=cert_id,
                revision=revision,
                cert_date=cert_date,
                product=str(row[5]) if row[5] else "",
                product_sn=str(row[6]) if row[6] else "",
                test_project=str(row[7]) if row[7] else "",
                project_name=str(row[8]) if row[8] else "",
                test_standard=str(row[9]) if row[9] else "",
                material=str(row[10]) if row[10] else "",
                specimen_id=str(row[11]) if row[11] else "",
                location_orientation=str(row[12]) if row[12] else "",
                temperature=str(row[13]) if row[13] else "",
                customer=str(row[14]) if row[14] else "",
                customer_order=str(row[15]) if row[15] else "",
                comment=str(row[16]) if row[16] else "",
                reported=str(row[17]).upper() == 'X' if row[17] else False,
                invoiced=str(row[18]).upper() == 'X' if row[18] else False
            )

            self.create_certificate(cert)
            imported += 1

        return imported
